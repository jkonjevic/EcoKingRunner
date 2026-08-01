from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import Browser, Error as PlaywrightError, Page, sync_playwright

from ecoking import stations as station_registry
from ecoking.stations import ExcelRow, Station


ROOT = Path(__file__).resolve().parent
ISO_DATE_FORMAT = "%Y-%m-%d"
WEBSITE_DATE_FORMAT = "%d/%m/%Y"
DEFAULT_TEMPLATE = "ECO KING BLANKO TABLICA.xlsx"

#: Report columns in the template, by header.
COLUMN_DAILY_M3 = 6
COLUMN_MAX_DAILY_M3 = 8
COLUMN_MIN_DAILY_M3 = 10
COLUMN_BATTERY = 13


def desktop_directory() -> Path | None:
    candidates = [
        Path.home() / "Desktop",
        Path(os.environ.get("OneDrive", "")) / "Desktop" if os.environ.get("OneDrive") else None,
    ]
    for candidate in candidates:
        if candidate and candidate.exists() and candidate.is_dir():
            return candidate
    return None


def store_report_on_desktop(report_path: Path) -> Path:
    desktop = desktop_directory()
    if desktop is None:
        logging.warning("Desktop folder was not found; keeping report at %s.", report_path)
        return report_path
    destination = desktop / report_path.name
    if report_path.resolve() != destination.resolve():
        shutil.copy2(report_path, destination)
    return destination


@dataclass(frozen=True)
class Measurement:
    daily_m3: float | None
    max_daily_m3: float | None
    min_daily_m3: float | None
    battery_voltage: str | None = None

    @property
    def daily_lps(self) -> float | None:
        return self.daily_m3 * 1000 / 86400 if self.daily_m3 is not None else None

    @property
    def max_daily_lps(self) -> float | None:
        return self.max_daily_m3 * 1000 / 900 if self.max_daily_m3 is not None else None

    @property
    def min_daily_lps(self) -> float | None:
        return self.min_daily_m3 * 1000 / 900 if self.min_daily_m3 is not None else None


@dataclass(frozen=True)
class StationJob:
    """A station paired with the template row its values belong in."""

    station: Station
    excel_row: int

    @property
    def key(self) -> tuple[str, str]:
        return self.station.key

    @property
    def label(self) -> str:
        return self.station.label


@dataclass(frozen=True)
class RunIssue:
    job: StationJob
    reason: str


@dataclass(frozen=True)
class RunResult:
    measurements: dict[tuple[str, str], Measurement]
    successes: list[StationJob]
    failures: list[RunIssue]
    no_data: list[RunIssue]


def env(name: str, default: str | None = None) -> str | None:
    return os.getenv(name) or os.getenv(name.lower()) or default


def as_bool(value: str | bool | None, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def one_line(text: object, limit: int | None = None) -> str:
    """Flatten an error message so it stays on a single, readable log line.

    Playwright errors carry multi-line call logs, and the UI colours the log
    line by line, so a wrapped message loses its colour after the first line.
    """
    collapsed = re.sub(r"\s+", " ", str(text)).strip()
    if limit and len(collapsed) > limit:
        return collapsed[: limit - 1].rstrip() + "…"
    return collapsed


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    use_color = not as_bool(os.getenv("NO_COLOR"), default=False)

    class ColorFormatter(logging.Formatter):
        COLORS = {
            logging.DEBUG: "\033[36m",
            logging.INFO: "\033[32m",
            logging.WARNING: "\033[33m",
            logging.ERROR: "\033[31m",
            logging.CRITICAL: "\033[1;31m",
        }
        RESET = "\033[0m"

        def format(self, record: logging.LogRecord) -> str:
            if use_color:
                color = self.COLORS.get(record.levelno, "")
                record.levelname = f"{color}{record.levelname}{self.RESET}"
                record.msg = f"{color}{record.msg}{self.RESET}" if record.levelno >= logging.WARNING else record.msg
            return super().format(record)

    handler = logging.StreamHandler()
    handler.setFormatter(ColorFormatter("%(asctime)s %(levelname)-17s %(message)s", datefmt="%H:%M:%S"))
    root = logging.getLogger()
    root.handlers.clear()
    root.setLevel(level)
    root.addHandler(handler)


def yesterday_date() -> datetime:
    return datetime.now() - timedelta(days=1)


def parse_selected_date(value: str | None) -> datetime:
    """Parse the UI/CLI date and reject future dates."""
    if not value:
        selected = yesterday_date()
    else:
        try:
            selected = next(
                datetime.strptime(value.strip(), date_format)
                for date_format in (ISO_DATE_FORMAT, WEBSITE_DATE_FORMAT)
                if _can_parse_date(value.strip(), date_format)
            )
        except StopIteration as exc:
            raise ValueError("Selected date must use YYYY-MM-DD or DD/MM/YYYY format.") from exc
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if selected > today:
        raise ValueError("Selected date cannot be in the future.")
    return selected


def _can_parse_date(value: str, date_format: str) -> bool:
    try:
        datetime.strptime(value, date_format)
        return True
    except ValueError:
        return False


def target_dates(selected_date: datetime) -> tuple[datetime, datetime]:
    """Return (Target_Date_Interval, Target_Date_Total)."""
    # Both diagrams use the date selected in the Python UI.
    return selected_date, selected_date


def load_template_rows(template_path: Path) -> list[ExcelRow]:
    return station_registry.load_excel_rows(template_path)


def load_stations(path: Path | None) -> list[Station]:
    """Read the station registry, upgrading the legacy file shape if needed."""
    resolved = station_registry.resolve_stations_path(path, ROOT)
    if not resolved.exists():
        raise FileNotFoundError(
            f"Station registry not found: {resolved}. Create it in the app or copy stations.json next to the script."
        )
    stations = station_registry.load_stations(resolved)
    logging.info("Loaded %s stations from %s", len(stations), resolved)
    return stations


def build_station_jobs(stations: Iterable[Station], rows: Iterable[ExcelRow]) -> list[StationJob]:
    """Pair every enabled station with its template row, skipping unmatched ones."""
    index = station_registry.index_excel_rows(rows)
    jobs: list[StationJob] = []
    for station in stations:
        if not station.enabled:
            logging.info("Station %r is disabled; skipping.", station.label)
            continue
        if not station.uredjaj:
            logging.warning("Station %r has no device name; skipping.", station.label)
            continue
        row = index.get(station.key)
        if row is None:
            logging.warning(
                "No template row for LOKACIJA=%r VODOMJER=%r; skipping %r.",
                station.lokacija,
                station.vodomjer,
                station.uredjaj,
            )
            continue
        jobs.append(StationJob(station=station, excel_row=row.row))
    return jobs


def clone_and_populate_template(
    template_path: Path,
    output_path: Path,
    selected_date: datetime,
    measurements: dict[tuple[str, str], Measurement],
    jobs: Iterable[StationJob],
) -> Path:
    """Copy the master workbook and fill in the scraped values.

    Only the value columns are written; the l/s formulas in G/I/K and every
    style in the template are left untouched.
    """
    if template_path.resolve() == output_path.resolve():
        raise RuntimeError(f"Output workbook must be different from {template_path.name}.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    workbook = load_workbook(output_path, data_only=False)
    worksheet = next(
        (
            candidate
            for candidate in workbook.worksheets
            if {str(cell.value).strip().upper() for cell in candidate[1] if cell.value}
            >= {"LOKACIJA", "VODOMJER"}
        ),
        None,
    )
    if worksheet is None:
        workbook.close()
        raise RuntimeError("Template is missing LOKACIJA and VODOMJER headers.")

    worksheet.title = selected_date.strftime(ISO_DATE_FORMAT)
    worksheet["N1"] = f"DATUM: {selected_date.strftime(ISO_DATE_FORMAT)}"

    row_by_key = {job.key: job.excel_row for job in jobs}
    mapped = 0
    unmapped: list[str] = []
    for key, measurement in measurements.items():
        excel_row = row_by_key.get(key)
        if excel_row is None:
            unmapped.append(" / ".join(key))
            continue
        worksheet.cell(row=excel_row, column=COLUMN_DAILY_M3).value = measurement.daily_m3
        worksheet.cell(row=excel_row, column=COLUMN_MAX_DAILY_M3).value = measurement.max_daily_m3
        worksheet.cell(row=excel_row, column=COLUMN_MIN_DAILY_M3).value = measurement.min_daily_m3
        worksheet.cell(row=excel_row, column=COLUMN_BATTERY).value = measurement.battery_voltage
        mapped += 1

    if unmapped:
        logging.warning("Could not place %s scraped station(s): %s", len(unmapped), ", ".join(unmapped))
    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    except AttributeError:
        logging.debug("Workbook calculation flags are not available in this openpyxl version.")
    workbook.save(output_path)
    workbook.close()
    logging.info(
        "Generated report %s for %s with %s mapped rows.",
        output_path,
        selected_date.strftime(ISO_DATE_FORMAT),
        mapped,
    )
    return output_path



def click_first_visible(page: Page, selectors: Iterable[str], label: str, timeout_ms: int = 10_000) -> bool:
    for selector in selectors:
        if not selector:
            continue
        locator = page.locator(selector).first
        try:
            locator.wait_for(state="visible", timeout=timeout_ms)
            logging.debug("Clicking %s with selector %s", label, selector)
            locator.click(timeout=timeout_ms)
            return True
        except PlaywrightError:
            continue
    return False


def fill_first_visible(page: Page, selectors: Iterable[str], value: str, label: str, timeout_ms: int = 10_000) -> bool:
    for selector in selectors:
        if not selector:
            continue
        locator = page.locator(selector).first
        try:
            locator.wait_for(state="visible", timeout=timeout_ms)
            logging.debug("Filling %s with selector %s", label, selector)
            locator.fill(value, timeout=timeout_ms)
            return True
        except PlaywrightError:
            continue
    return False


def fill_open_dropdown_search(page: Page, selectors: Iterable[str], value: str) -> bool:
    selector_list = [selector for selector in selectors if selector]
    filled_selector = page.evaluate(
        """
        ({selectors, value}) => {
          const visible = el => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0;
          };
          const setValue = (el, text) => {
            el.focus();
            const proto = el instanceof HTMLTextAreaElement
              ? window.HTMLTextAreaElement.prototype
              : window.HTMLInputElement.prototype;
            const valueSetter = Object.getOwnPropertyDescriptor(proto, 'value')?.set;
            if (valueSetter) {
              valueSetter.call(el, '');
              el.dispatchEvent(new Event('input', {bubbles: true}));
              valueSetter.call(el, text);
            } else {
              el.value = text;
            }
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
            el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true, key: text.slice(-1) || ' '}));
          };

          for (const selector of selectors) {
            let elements = [];
            try {
              elements = Array.from(document.querySelectorAll(selector));
            } catch {
              continue;
            }
            const input = elements.find(el =>
              visible(el)
              && (el instanceof HTMLInputElement || el instanceof HTMLTextAreaElement)
            );
            if (input) {
              setValue(input, value);
              return selector;
            }
          }
          return null;
        }
        """,
        {"selectors": selector_list, "value": value},
    )
    if filled_selector:
        logging.debug("Filling location search with visible dropdown input %s", filled_selector)
        return True

    try:
        active_tag = page.evaluate("document.activeElement ? document.activeElement.tagName.toLowerCase() : ''")
        if active_tag in {"input", "textarea"}:
            page.keyboard.press("Control+A")
            page.keyboard.type(value, delay=5)
            logging.debug("Typed location search into focused %s element", active_tag)
            return True
    except PlaywrightError:
        pass

    return False


def type_open_dropdown_search(page: Page, selectors: Iterable[str], value: str) -> bool:
    selector_list = [selector for selector in selectors if selector]
    focused_selector = page.evaluate(
        """
        (selectors) => {
          const visible = el => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0;
          };
          for (const selector of selectors) {
            let elements = [];
            try {
              elements = Array.from(document.querySelectorAll(selector));
            } catch {
              continue;
            }
            const input = elements.find(el =>
              visible(el)
              && (el instanceof HTMLInputElement || el instanceof HTMLTextAreaElement)
            );
            if (input) {
              input.focus();
              return selector;
            }
          }
          return null;
        }
        """,
        selector_list,
    )
    if not focused_selector:
        return False

    type_delay_ms = int(env("SEARCH_TYPE_DELAY_MS", "10") or "10")
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    page.keyboard.type(value, delay=type_delay_ms)
    logging.debug("Retyped location search with keyboard into %s", focused_selector)
    return True


#: Marks the dropdown entries a read has seen, so the follow-up click targets
#: exactly the element that produced the text -- no index bookkeeping.
_OPTION_TAG = "ecokingOption"
_OPTION_ATTR = "data-ecoking-option"

_READ_DROPDOWN_OPTIONS_JS = """
(tag) => {
  const visible = el => {
    if (!el) return false;
    const style = getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.visibility !== 'hidden'
      && style.display !== 'none'
      && rect.width > 0
      && rect.height > 0;
  };
  const normalize = text => (text || '')
    .normalize('NFD')
    .replace(/[\u0300-\u036f]/g, '')
    .toLowerCase()
    .replace(/[^a-z0-9]+/g, ' ')
    .trim();
  const skip = new Set([
    'home', 'device list', 'devices', 'user settings',
    'log out', 'logout', 'login', 'menu', 'settings'
  ]);
  const selector = ".dropdown-menu a, .dropdown-menu button, [role=listbox] [role=option], [role=option], option";
  for (const stale of document.querySelectorAll(`[data-${tag}]`)) delete stale.dataset[tag];

  const seen = new Set();
  const options = [];
  for (const element of document.querySelectorAll(selector)) {
    if (!visible(element)) continue;
    const text = (element.innerText || element.textContent || '').replace(/\s+/g, ' ').trim();
    if (text.length < 3) continue;
    const key = normalize(text);
    if (!key || skip.has(key) || seen.has(key)) continue;
    if (/no\s+(results?|matches?)|not\s+found|nema\s+rezultata/i.test(text)) continue;
    seen.add(key);
    element.dataset[tag] = String(options.length);
    options.push({index: options.length, text});
  }
  return options;
}
"""


def read_dropdown_options(page: Page) -> list[dict[str, Any]]:
    """Return every visible dropdown entry, tagged so it can be clicked later."""
    options = page.evaluate(_READ_DROPDOWN_OPTIONS_JS, _OPTION_TAG)
    return list(options or [])


def click_dropdown_option(page: Page, index: int) -> None:
    page.locator(f'[{_OPTION_ATTR}="{index}"]').first.click(timeout=5_000)


def wait_for_dropdown_options(page: Page, wait_ms: int | None = None) -> list[dict[str, Any]]:
    wait_ms = wait_ms if wait_ms is not None else int(env("SEARCH_RESULTS_WAIT_MS", "2000") or "2000")
    deadline = time.monotonic() + (wait_ms / 1000)
    options: list[dict[str, Any]] = []
    while True:
        options = read_dropdown_options(page)
        if options or time.monotonic() >= deadline:
            return options
        page.wait_for_timeout(150)


def choose_device_option(page: Page, uredjaj: str, wait_ms: int | None = None) -> str:
    """Click the dropdown entry that matches ``uredjaj`` and return its text.

    Entries are ranked against the stored label after their serial number and
    city prefix are stripped, so the registry never has to carry a serial. A tie
    at the top means the label is not specific enough to pick a device.
    """
    options = wait_for_dropdown_options(page, wait_ms)
    if not options:
        raise RuntimeError(f"No dropdown results after searching {uredjaj!r}.")

    scored = [
        (station_registry.score_device_match(uredjaj, station_registry.device_label(option["text"])), option)
        for option in options
    ]
    matches = [(score, option) for score, option in scored if score > 0]
    logging.debug("Dropdown options for %r: %s", uredjaj, [option["text"] for option in options])

    if not matches:
        raise RuntimeError(
            f"No dropdown result matches {uredjaj!r}. Visible: {[option['text'] for option in options[:10]]}"
        )

    best_score = max(score for score, _ in matches)
    best = [option for score, option in matches if score == best_score]
    if len(best) > 1:
        raise RuntimeError(
            f"Device name {uredjaj!r} matches {len(best)} devices: {[option['text'] for option in best]}. "
            "Make the name in the station list more specific."
        )

    chosen = best[0]
    logging.info("Choosing device: %s", chosen["text"])
    click_dropdown_option(page, int(chosen["index"]))
    page.wait_for_timeout(250)
    return str(chosen["text"])

def click_interval_option(page: Page, label: str) -> bool:
    clicked = page.evaluate(
        """
        (label) => {
          const visible = el => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0;
          };
          const normalize = text => (text || '').replace(/\\s+/g, ' ').trim();
          const candidates = Array.from(document.querySelectorAll("button, a, [role=button], .dropdown-menu a"))
            .filter(visible)
            .filter(el => normalize(el.innerText || el.textContent) === label)
            .map(el => {
              let score = 0;
              const attrs = Array.from(el.attributes || []).map(attr => `${attr.name}=${attr.value}`).join(' ');
              const cls = String(el.className || '');
              if (/changeTimeSlice/i.test(attrs)) score += 100;
              if (el.tagName.toLowerCase() === 'button') score += 20;
              if (/dropdown-toggle/i.test(cls) || /uib-dropdown-toggle/i.test(attrs)) score -= 50;
              if (/btn-primary/i.test(cls)) score += 10;
              return {el, score};
            })
            .sort((a, b) => b.score - a.score);
          if (!candidates.length) return false;
          candidates[0].el.click();
          return true;
        }
        """,
        label,
    )
    if clicked:
        logging.debug("Clicked interval option %r", label)
    return bool(clicked)


def click_device_dropdown(page: Page) -> bool:
    configured = env("SEARCH_TOGGLE_SELECTOR")
    if configured:
        return click_first_visible(page, [configured], "configured device dropdown")

    clicked_precise = page.evaluate(
        """
        (() => {
          const visible = el => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0;
          };
          const candidates = Array.from(document.querySelectorAll("button.btn-primary.dropdown-toggle, .btn-primary.dropdown-toggle"))
            .filter(visible)
            .map(el => ({el, rect: el.getBoundingClientRect()}))
            .filter(item =>
              item.rect.left >= 0
              && item.rect.left < 90
              && item.rect.top > 40
              && item.rect.top < 160
              && item.rect.width >= 20
              && item.rect.width <= 70
              && item.rect.height >= 20
              && item.rect.height <= 55
            )
            .sort((a, b) => (a.rect.left - b.rect.left) || (a.rect.top - b.rect.top));
          if (!candidates.length) return false;
          candidates[0].el.click();
          return true;
        })()
        """
    )
    if clicked_precise:
        logging.debug("Clicked device dropdown using precise top-left selector")
        return True

    clicked = page.evaluate(
        """
        (() => {
          const visible = el => {
            if (!el) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden'
              && style.display !== 'none'
              && rect.width > 0
              && rect.height > 0;
          };
          const score = el => {
            const rect = el.getBoundingClientRect();
            const text = (el.innerText || el.getAttribute('aria-label') || el.getAttribute('title') || '').trim();
            const html = el.innerHTML || '';
            const cls = String(el.className || '');
            let value = 0;
            if (rect.left < 80 && rect.top > 45 && rect.top < 140) value += 100;
            if (rect.width >= 20 && rect.width <= 55 && rect.height >= 20 && rect.height <= 45) value += 30;
            if (/btn-primary|dropdown-toggle|select2-choice|multiselect/i.test(cls)) value += 20;
            if (/caret|angle-down|chevron-down|fa-caret-down|fa-angle-down|▼|▾|⌄|⌵/i.test(html + ' ' + text)) value += 20;
            if (rect.top < 45 || rect.left > 160) value -= 80;
            if (/navbar|sidebar|hamburger|menu|bars|toggle/i.test(cls + ' ' + html + ' ' + text)) value -= 50;
            return value;
          };
          const candidates = Array.from(document.querySelectorAll("button, a, [role=button]"))
            .filter(visible)
            .map(el => ({el, value: score(el)}))
            .filter(item => item.value > 0)
            .sort((a, b) => b.value - a.value);
          if (!candidates.length) return false;
          candidates[0].el.click();
          return true;
        })()
        """
    )
    if clicked:
        logging.debug("Clicked device dropdown using top-left content heuristic")
        return True

    return click_first_visible(
        page,
        [
            "button.btn-primary.dropdown-toggle",
            ".btn-primary.dropdown-toggle",
            "button:has(.fa-angle-down)",
            "button:has(.fa-caret-down)",
            "button:has-text('▼')",
            "button:has-text('▾')",
        ],
        "device dropdown",
    )


def select_website_date(page: Page, website_target_date: datetime) -> None:
    """Choose the site's date (+1 day from the requested reporting date)."""
    target_display = website_target_date.strftime(WEBSITE_DATE_FORMAT)
    target_day = str(website_target_date.day)
    picker_selector = env("WEBSITE_DATE_PICKER_SELECTOR")

    if picker_selector:
        if not click_first_visible(page, [picker_selector], "website date picker"):
            logging.debug("Could not click WEBSITE_DATE_PICKER_SELECTOR; trying the inline top-right date field.")
    else:
        # This is the site's Angular UI Bootstrap control. Keep it ahead of
        # positional heuristics so the hamburger/menu buttons can never win.
        clicked = click_first_visible(
            page,
            [
                "div[uib-tooltip='Choose Date...'][ng-click='chooseDate()']",
                "[uib-tooltip='Choose Date...']",
            ],
            "website date picker (Choose Date...)",
            timeout_ms=3_000,
        )
        if not clicked:
            clicked = page.evaluate(
            """
            (() => {
              const visible = el => {
                const s = getComputedStyle(el), r = el.getBoundingClientRect();
                return s.visibility !== 'hidden' && s.display !== 'none' && r.width > 0 && r.height > 0;
              };
              const candidates = Array.from(document.querySelectorAll('button, a, [role=button], input'))
                .filter(visible)
                .map(el => {
                  const r = el.getBoundingClientRect();
                  const text = [el.getAttribute('title'), el.getAttribute('aria-label'),
                    el.getAttribute('data-original-title'), el.innerText, el.textContent]
                    .filter(Boolean).join(' ');
                  let score = /choose date/i.test(text) ? 1000 : 0;
                  if (r.left > window.innerWidth * .65) score += 100;
                  if (r.top < window.innerHeight * .35) score += 20;
                  return {el, score};
                }).filter(item => item.score > 0).sort((a, b) => b.score - a.score);
              if (!candidates.length) return false;
              candidates[0].el.click();
              return true;
            })()
            """
            )
        if not clicked:
            logging.debug("Date picker tooltip button was not directly identified; trying the inline top-right date field.")

    # Some versions render the date selector inline in the top-right toolbar instead
    # of opening a modal. The field is identifiable by its DD/MM/YYYY value.
    def fill_inline_date_field() -> bool:
        selector = env("WEBSITE_DATE_INPUT_SELECTOR")
        if selector:
            candidates = page.locator(selector)
        else:
            candidates = page.locator("input, [role='textbox']")

        def navigate_with_adjacent_buttons(candidate: Any) -> bool:
            try:
                field_box = candidate.bounding_box()
                if not field_box:
                    return False
                buttons: list[tuple[float, Any]] = []
                toolbar_buttons = page.locator("button, [role='button']")
                for button_index in range(toolbar_buttons.count()):
                    button = toolbar_buttons.nth(button_index)
                    if not button.is_visible(timeout=300):
                        continue
                    button_box = button.bounding_box()
                    if not button_box or abs(button_box[1] - field_box[1]) > max(field_box[3], 50):
                        continue
                    if button_box[0] + button_box[2] <= field_box[0]:
                        buttons.append((field_box[0] - (button_box[0] + button_box[2]), button))
                    elif button_box[0] >= field_box[0] + field_box[2]:
                        buttons.append((button_box[0] - (field_box[0] + field_box[2]), button))
                left_buttons = sorted((distance, button) for distance, button in buttons if button.bounding_box()[0] < field_box[0])
                right_buttons = sorted((distance, button) for distance, button in buttons if button.bounding_box()[0] >= field_box[0] + field_box[2])
                previous = left_buttons[0][1] if left_buttons else None
                following = right_buttons[0][1] if right_buttons else None
                if not previous or not following:
                    return False
                for _ in range(400):
                    value = candidate.input_value().strip()
                    parsed = None
                    for value_format in (WEBSITE_DATE_FORMAT, ISO_DATE_FORMAT):
                        try:
                            parsed = datetime.strptime(value, value_format)
                            break
                        except ValueError:
                            continue
                    if parsed is None:
                        return False
                    if parsed.date() == website_target_date.date():
                        logging.info("Reached inline website date %s using toolbar navigation", target_display)
                        return True
                    control = previous if parsed.date() > website_target_date.date() else following
                    control.click()
                    page.wait_for_timeout(100)
                return False
            except (PlaywrightError, TypeError, ValueError):
                return False

        for index in range(candidates.count()):
            candidate = candidates.nth(index)
            try:
                if not candidate.is_visible(timeout=500):
                    continue
                box = candidate.bounding_box()
                value = candidate.input_value()
                viewport_width = (page.viewport_size or {"width": 1440})["width"]
                if not box or box[0] < viewport_width * 0.55:
                    continue
                if not selector and not (
                    re.fullmatch(r"\d{2}/\d{2}/\d{4}", value.strip())
                    or re.fullmatch(r"\d{4}-\d{2}-\d{2}", value.strip())
                ):
                    continue
                candidate.click(timeout=1_000)
                try:
                    input_type = candidate.get_attribute("type")
                    candidate.fill(target_iso if input_type == "date" else target_display, timeout=1_000)
                except PlaywrightError:
                    if navigate_with_adjacent_buttons(candidate):
                        return True
                    page.evaluate(
                        """([element, value]) => {
                          const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
                          setter.call(element, value);
                          element.dispatchEvent(new Event('input', {bubbles: true}));
                          element.dispatchEvent(new Event('change', {bubbles: true}));
                        }""",
                        [candidate.element_handle(), target_iso if input_type == "date" else target_display],
                    )
                candidate.press("Enter")
                candidate.blur()
                logging.info("Entered inline website date %s in the top-right date control", target_display)
                return True
            except (PlaywrightError, TypeError, ValueError):
                continue
        return False

    dialog_selectors = [
        env("WEBSITE_DATE_DIALOG_SELECTOR"),
        "[role='dialog']",
        ".modal:visible",
        ".ui-dialog:visible",
        ".datepicker:visible",
    ]
    dialog = None
    for selector in dialog_selectors:
        if not selector:
            continue
        candidate = page.locator(selector).last
        try:
            candidate.wait_for(state="visible", timeout=3_000)
            dialog = candidate
            break
        except PlaywrightError:
            continue
    if dialog is None:
        if fill_inline_date_field():
            return
        raise RuntimeError("Date picker dialog did not open and no top-right DD/MM/YYYY date field was found.")

    date_input_selector = env("WEBSITE_DATE_INPUT_SELECTOR")
    if date_input_selector:
        date_input = dialog.locator(date_input_selector).first
        date_input.fill(target_display)
        logging.debug("Entered website date %s using WEBSITE_DATE_INPUT_SELECTOR", target_display)

    # Prefer site-specific selectors, then common Bootstrap/Angular datepicker labels.
    prev = env("WEBSITE_DATE_PREV_SELECTOR")
    next_ = env("WEBSITE_DATE_NEXT_SELECTOR")
    if not prev or not next_:
        prev = prev or "button[aria-label*='Previous'], button[title*='Previous'], button[title*='previous'], button[ng-click='move(-1)']"
        next_ = next_ or "button[aria-label*='Next'], button[title*='Next'], button[title*='next'], button[ng-click='move(1)']"

    # The calendar opens on the month of the date that is currently applied on
    # the site, not on the current month, so the number of steps to take can
    # only be derived from the header the widget is actually showing. Read it
    # each iteration and step until it matches, rather than assuming a
    # starting point.
    month_titles = {
        "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
        "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12,
    }

    def displayed_month() -> tuple[int, int] | None:
        title = page.evaluate(
            """() => {
              const el = document.querySelector('.uib-title, [role=dialog] .uib-title');
              return el ? (el.innerText || '').trim() : '';
            }"""
        )
        match = re.search(r"([A-Za-z]+)\s+(\d{4})", str(title or ""))
        if not match or match.group(1).lower() not in month_titles:
            return None
        return int(match.group(2)), month_titles[match.group(1).lower()]

    target_month = (website_target_date.year, website_target_date.month)
    for _ in range(120):
        shown = displayed_month()
        if shown is None or shown == target_month:
            break
        step_forward = shown < target_month
        step_selector = next_ if step_forward else prev
        step_label = "next calendar month" if step_forward else "previous calendar month"
        if not click_first_visible(page, [step_selector], step_label, timeout_ms=1_000):
            raise RuntimeError("Could not navigate to the requested calendar month. Set WEBSITE_DATE_NEXT_SELECTOR/WEBSITE_DATE_PREV_SELECTOR.")
        page.wait_for_timeout(100)

    # Day cells are zero-padded ("01"), the grid pads each month with the
    # neighbouring months' days (marked text-muted on the inner span), and a
    # week-number column carries bare numbers that would also match a day.
    # Restrict to real day cells and accept the padded form.
    day_selector = env("WEBSITE_DATE_DAY_SELECTOR")
    day_pattern = re.compile(rf"^0*{re.escape(target_day)}$")
    if day_selector:
        candidates = dialog.locator(day_selector).filter(has_text=day_pattern)
    else:
        candidates = dialog.locator("td.uib-day button")
        if not candidates.count():
            candidates = dialog.locator("td[role='gridcell'] button, td button")

    day = None
    for index in range(candidates.count()):
        candidate = candidates.nth(index)
        try:
            if not day_pattern.match((candidate.inner_text() or "").strip()):
                continue
            if candidate.locator("span.text-muted").count():
                continue
            if not candidate.is_enabled(timeout=500):
                continue
        except (PlaywrightError, TypeError):
            continue
        day = candidate
        break

    if day is None:
        raise RuntimeError(
            f"Could not find a selectable cell for website date {target_display} "
            f"in the displayed month ({candidates.count()} day cell(s) scanned). Set WEBSITE_DATE_DAY_SELECTOR."
        )
    try:
        day.click(timeout=5_000)
    except PlaywrightError as exc:
        raise RuntimeError(f"Could not select website date {target_display}. Set WEBSITE_DATE_DAY_SELECTOR.") from exc

    ok_selector = env("WEBSITE_DATE_OK_SELECTOR") or "button:has-text('OK'), button:has-text('Ok'), button:has-text('Apply')"
    if not click_first_visible(page, [ok_selector], "calendar OK button", timeout_ms=5_000):
        raise RuntimeError("Could not confirm the date picker selection. Set WEBSITE_DATE_OK_SELECTOR.")
    logging.info("Selected website date %s", target_display)


def select_date_for_metric(page: Page, metric_name: str, target_date: datetime) -> None:
    """Apply and log the date required by one diagram/metric."""
    logging.info("DATE APPLY | metric=%s | website date=%s", metric_name, target_date.strftime(WEBSITE_DATE_FORMAT))
    select_website_date(page, target_date)


def login(page: Page, base_url: str, email: str, password: str) -> None:
    logging.info("Opening %s", base_url)
    page.goto(base_url, wait_until="domcontentloaded")
    page.wait_for_load_state("domcontentloaded", timeout=30_000)

    email_selectors = [
        env("LOGIN_EMAIL_SELECTOR"),
        "input[type='email']",
        "input[name='email']",
        "input[name='username']",
        "#email",
        "#username",
        "input[type='text']",
    ]
    password_selectors = [
        env("LOGIN_PASSWORD_SELECTOR"),
        "input[type='password']",
        "input[name='password']",
        "#password",
    ]
    submit_selectors = [
        env("LOGIN_SUBMIT_SELECTOR"),
        "button[type='submit']",
        "input[type='submit']",
        "button:has-text('Login')",
        "button:has-text('Log in')",
        "button:has-text('Sign in')",
        "button:has-text('Prijava')",
    ]

    if fill_first_visible(page, email_selectors, email, "email"):
        click_first_visible(page, submit_selectors, "login next/submit", timeout_ms=3_000)
        try:
            page.wait_for_timeout(1_000)
        except PlaywrightError:
            pass

    if fill_first_visible(page, password_selectors, password, "password"):
        if not click_first_visible(page, submit_selectors, "login submit", timeout_ms=5_000):
            page.keyboard.press("Enter")
        page.wait_for_timeout(1_000)
        logging.info("Login flow submitted")
        return

    wait_seconds = int(env("WAIT_FOR_LOGIN_SECONDS", "0") or "0")
    if wait_seconds > 0:
        logging.warning("Login fields were not detected. Waiting %s seconds for manual login.", wait_seconds)
        page.wait_for_timeout(wait_seconds * 1000)
        return

    raise RuntimeError("Could not complete login. Add LOGIN_* selectors or set WAIT_FOR_LOGIN_SECONDS.")


SEARCH_INPUT_SELECTORS = [
    ".dropdown-menu input",
    ".dropdown-menu textarea",
    ".dropdown-menu .form-control",
    ".dropdown.open input",
    ".open > .dropdown-menu input",
    ".uib-dropdown-menu input",
    "[role='listbox'] input",
    ".select2-search__field",
    ".select2-search input",
    "input[type='search']",
]


def search_input_selectors() -> list[str]:
    configured = env("SEARCH_INPUT_SELECTOR")
    return [configured, *SEARCH_INPUT_SELECTORS] if configured else list(SEARCH_INPUT_SELECTORS)


def open_device_dropdown(page: Page, query: str) -> None:
    """Open the device dropdown and type ``query`` into its search box."""
    try:
        page.keyboard.press("Escape")
        page.wait_for_timeout(100)
    except PlaywrightError:
        pass

    if not click_device_dropdown(page):
        raise RuntimeError("Could not find the blue dropdown/search button. Set SEARCH_TOGGLE_SELECTOR.")

    page.wait_for_timeout(150)
    if not query:
        return

    selectors = search_input_selectors()
    input_wait_ms = int(env("SEARCH_INPUT_WAIT_MS", "500") or "500")
    if not fill_open_dropdown_search(page, selectors, query):
        if not fill_first_visible(page, selectors, query, "device search", timeout_ms=input_wait_ms):
            raise RuntimeError("Could not find the opened dropdown search input. Set SEARCH_INPUT_SELECTOR.")
    page.wait_for_timeout(250)


def select_device(page: Page, uredjaj: str) -> None:
    """Search for ``uredjaj`` on the site and open its diagram page.

    Each query is typed through the DOM first; if the site's filter ignores that,
    the same query is retyped on the keyboard before giving up and falling back
    to a shorter query.
    """
    logging.info("Searching device: %s", uredjaj)
    queries = station_registry.search_queries(uredjaj)
    errors: list[str] = []

    for query in queries:
        try:
            open_device_dropdown(page, query)
            try:
                choose_device_option(page, uredjaj)
            except RuntimeError as exc:
                if "No dropdown results" not in str(exc):
                    raise
                logging.warning("No dropdown results for %r after DOM fill. Retrying with keyboard typing.", query)
                if not type_open_dropdown_search(page, search_input_selectors(), query):
                    raise
                page.wait_for_timeout(500)
                choose_device_option(page, uredjaj)
            if query != uredjaj:
                logging.info("Selected %r using shorter query %r", uredjaj, query)
            return
        except Exception as exc:
            errors.append(f"{query!r}: {exc}")
            if query != queries[-1]:
                logging.warning("Search query %r failed for %r. Trying a shorter query.", query, uredjaj)
                continue
            raise RuntimeError(f"Could not select device {uredjaj!r}. Attempts: {' | '.join(errors)}") from exc


def wait_for_chart_data(page: Page, timeout_ms: int | None = None) -> None:
    timeout_ms = timeout_ms or int(env("CHART_WAIT_MS", "5000") or "5000")
    page.wait_for_function(
        """
        () => {
          if (window.Highcharts && Array.isArray(window.Highcharts.charts)) {
            const hasHighchartsData = window.Highcharts.charts
              .filter(Boolean)
              .some(chart => (chart.series || []).some(s => (s.points || []).some(p => p.y !== null && p.y !== undefined)));
            if (hasHighchartsData) return true;
          }
          if (window.Chart) {
            const charts = window.Chart.instances
              ? Object.values(window.Chart.instances)
              : (window.Chart.getChart ? Array.from(document.querySelectorAll('canvas')).map(c => window.Chart.getChart(c)).filter(Boolean) : []);
            const hasChartJsData = charts.some(chart =>
              chart.data && (chart.data.datasets || []).some(dataset => (dataset.data || []).length)
            );
            if (hasChartJsData) return true;
          }
          return Array.from(document.querySelectorAll('table tr')).some(tr => tr.querySelectorAll('td,th').length > 1);
        }
        """,
        timeout=timeout_ms,
    )


def read_battery_voltage(page: Page, timeout_ms: int | None = None) -> str | None:
    """Read the visible Battery level stat for the currently selected station."""
    timeout_ms = timeout_ms or int(env("BATTERY_WAIT_MS", "5000") or "5000")
    selector = env("BATTERY_SELECTOR") or "span[uib-tooltip='Battery level']"
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        locator = page.locator(selector)
        for index in range(locator.count()):
            candidate = locator.nth(index)
            try:
                if not candidate.is_visible(timeout=300):
                    continue
                text = re.sub(r"\s+", " ", candidate.inner_text()).strip()
                match = re.search(r"(\d+(?:[.,]\d+)?)\s*V\b", text, flags=re.IGNORECASE)
                if match:
                    voltage = match.group(1).replace(",", ".")
                    value = f"{voltage} V"
                    logging.info("Read battery level: %s", value)
                    return value
            except PlaywrightError:
                continue
        page.wait_for_timeout(100)
    logging.warning("Battery level component was not found for the selected station.")
    return None


def select_interval(page: Page, label: str) -> None:
    logging.info("Selecting interval: %s", label)
    configured = env("INTERVAL_SELECTOR")
    if configured:
        control = page.locator(configured).first
        control.wait_for(state="visible", timeout=10_000)
        tag = control.evaluate("el => el.tagName.toLowerCase()")
        if tag == "select":
            control.select_option(label=label)
        else:
            control.click()
            if not click_interval_option(page, label):
                raise RuntimeError(f"Could not choose interval {label!r} using INTERVAL_SELECTOR.")
    else:
        select = page.locator("select").filter(has_text=re.compile(re.escape(label), re.I)).first
        try:
            select.select_option(label=label, timeout=5_000)
        except PlaywrightError:
            if click_interval_option(page, label):
                wait_for_chart_data(page)
                return
            if click_first_visible(page, [f"button:has-text('{label}')"], "interval button", timeout_ms=5_000):
                wait_for_chart_data(page)
                return
            if not click_first_visible(page, ["[role='combobox']", ".dropdown-toggle"], "interval dropdown", timeout_ms=5_000):
                raise RuntimeError("Could not find interval selector. Set INTERVAL_SELECTOR.")
            if not click_interval_option(page, label):
                raise RuntimeError(f"Could not choose interval {label!r}.")

    wait_for_chart_data(page)


def parse_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"[^\d,.\-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def extract_chart_payload(page: Page) -> dict[str, Any]:
    return page.evaluate(
        """
        () => {
          const payload = {highcharts: [], chartjs: [], tables: []};

          if (window.Highcharts && Array.isArray(window.Highcharts.charts)) {
            for (const chart of window.Highcharts.charts.filter(Boolean)) {
              const series = (chart.series || []).map(s => ({
                name: s.name,
                type: s.type,
                points: (s.points || []).map(p => ({
                  x: p.x,
                  y: p.y,
                  category: p.category,
                  name: p.name,
                  // This is the date as Highcharts formats the X axis. It is
                  // more authoritative than converting p.x in Python because
                  // it uses the chart's own timezone configuration.
                  date: typeof p.x === 'number' && window.Highcharts.dateFormat
                    ? window.Highcharts.dateFormat('%Y-%m-%d', p.x)
                    : null
                }))
              }));
              payload.highcharts.push({series});
            }
          }

          if (window.Chart) {
            const charts = window.Chart.instances
              ? Object.values(window.Chart.instances)
              : (window.Chart.getChart ? Array.from(document.querySelectorAll('canvas')).map(c => window.Chart.getChart(c)).filter(Boolean) : []);
            for (const chart of charts) {
              payload.chartjs.push({
                labels: chart.data && chart.data.labels || [],
                datasets: chart.data && chart.data.datasets || []
              });
            }
          }

          for (const table of document.querySelectorAll('table')) {
            const rows = Array.from(table.querySelectorAll('tr')).map(tr =>
              Array.from(tr.querySelectorAll('th,td')).map(td => td.innerText.trim())
            ).filter(row => row.length);
            if (rows.length) payload.tables.push(rows);
          }
          return payload;
        }
        """
    )


def numbers_from_chart_payload(payload: dict[str, Any]) -> list[float]:
    values: list[float] = []
    for chart in payload.get("highcharts", []):
        for series in chart.get("series", []):
            for point in series.get("points", []):
                value = parse_float(point.get("y"))
                if value is not None:
                    values.append(value)
    for chart in payload.get("chartjs", []):
        for dataset in chart.get("datasets", []):
            for raw in dataset.get("data", []):
                value = parse_float(raw.get("y") if isinstance(raw, dict) else raw)
                if value is not None:
                    values.append(value)
    return values


def latest_value_from_payload(payload: dict[str, Any]) -> float | None:
    highcharts_points: list[tuple[float, float]] = []
    for chart in payload.get("highcharts", []):
        for series in chart.get("series", []):
            for point in series.get("points", []):
                y = parse_float(point.get("y"))
                x = parse_float(point.get("x"))
                if y is not None and x is not None:
                    highcharts_points.append((x, y))
    if highcharts_points:
        return max(highcharts_points, key=lambda item: item[0])[1]

    chartjs_values: list[float] = []
    for chart in payload.get("chartjs", []):
        for dataset in chart.get("datasets", []):
            for raw in dataset.get("data", []):
                value = parse_float(raw.get("y") if isinstance(raw, dict) else raw)
                if value is not None:
                    chartjs_values.append(value)
    if chartjs_values:
        return chartjs_values[-1]

    for table in payload.get("tables", []):
        for row in reversed(table):
            values = [parse_float(cell) for cell in row]
            values = [value for value in values if value is not None]
            if values:
                return values[-1]
    return None


def chart_point_date(value: Any, target_year: int | None = None) -> date | None:
    """Convert a chart point/category value to a calendar date when possible."""
    if isinstance(value, (int, float)):
        # Highcharts uses milliseconds since the Unix epoch for point.x.
        timestamp = float(value)
        if abs(timestamp) > 10_000_000_000:
            timestamp /= 1000
        try:
            # Highcharts renders datetime-axis labels in the browser's local
            # timezone.  Using UTC here shifts bars around midnight to the
            # previous calendar day (e.g. the visible 24 Jul bar becomes 23
            # Jul), so match the same local date the user sees.
            return datetime.fromtimestamp(timestamp).date()
        except (OverflowError, OSError, ValueError):
            return None

    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return None

    # Prefer explicit numeric dates; chart labels commonly include a time too.
    patterns = (
        (r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", lambda m: (int(m.group(1)), int(m.group(2)), int(m.group(3)))),
        (r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})", lambda m: (int(m.group(3)), int(m.group(2)), int(m.group(1)))),
    )
    for pattern, parts in patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return date(*parts(match))
            except ValueError:
                return None

    # Some chart labels omit the year (for example, "23/07").  Only use this
    # when the requested year is known, so an ambiguous label is not guessed.
    if target_year is not None:
        match = re.search(r"\b(\d{1,2})[-/.](\d{1,2})\b", text)
        if match:
            try:
                return date(target_year, int(match.group(2)), int(match.group(1)))
            except ValueError:
                return None
    return None


def value_for_date_from_payload(payload: dict[str, Any], target_date: datetime) -> float | None:
    """Return the bar whose chart date equals target_date, never simply the last bar."""
    wanted = target_date.date()

    highcharts_matches: list[tuple[int, float]] = []
    for chart in payload.get("highcharts", []):
        for series in chart.get("series", []):
            for point in series.get("points", []):
                value = parse_float(point.get("y"))
                if value is None:
                    continue
                point_dates = {
                    chart_point_date(point.get("x"), wanted.year),
                    chart_point_date(point.get("category"), wanted.year),
                    chart_point_date(point.get("name"), wanted.year),
                    chart_point_date(point.get("date"), wanted.year),
                }
                if wanted in point_dates:
                    series_name = str(series.get("name") or "").lower()
                    series_type = str(series.get("type") or "").lower()
                    # Prefer the actual usage bar series when the page has
                    # more than one Highcharts series with dated points.
                    priority = 0 if series_type in {"column", "bar"} else 10
                    if "usage" in series_name or "potro" in series_name or "cubic" in series_name:
                        priority -= 1
                    highcharts_matches.append((priority, value))

    if highcharts_matches:
        return min(highcharts_matches, key=lambda item: item[0])[1]

    for chart in payload.get("chartjs", []):
        labels = chart.get("labels", [])
        for dataset in chart.get("datasets", []):
            for index, raw in enumerate(dataset.get("data", [])):
                value = parse_float(raw.get("y") if isinstance(raw, dict) else raw)
                if value is None:
                    continue
                point_dates = {chart_point_date(labels[index], wanted.year)} if index < len(labels) else set()
                if isinstance(raw, dict):
                    point_dates.update({
                        chart_point_date(raw.get("x"), wanted.year),
                        chart_point_date(raw.get("label"), wanted.year),
                    })
                if wanted in point_dates:
                    return value

    return None


def save_debug_artifacts(page: Page, debug_dir: Path, location: str, suffix: str) -> None:
    debug_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", location).strip("_") or "location"
    stem = debug_dir / f"{safe_name}_{suffix}_{int(time.time())}"
    page.screenshot(path=str(stem.with_suffix(".png")), full_page=True)
    stem.with_suffix(".html").write_text(page.content(), encoding="utf-8")
    logging.warning("Saved debug artifacts: %s.png and %s.html", stem, stem)


def scrape_measurement(
    page: Page,
    location: str,
    target_date_interval: datetime,
    target_date_total: datetime,
    debug_dir: Path,
) -> Measurement:
    select_device(page, location)
    battery_voltage = read_battery_voltage(page)

    # The 30-day diagram is opened at the date selected in the Python UI and
    # its last bar is the requested daily total.
    select_date_for_metric(page, "UKUPNA DNEVNA POTROŠNJA (m3) / 30-day", target_date_total)
    select_interval(page, env("INTERVAL_30_DAYS_LABEL", "30 days") or "30 days")
    payload_30_days = extract_chart_payload(page)
    daily = latest_value_from_payload(payload_30_days)
    if daily is None:
        save_debug_artifacts(page, debug_dir, location, "30days_no_selected_date")
        raise RuntimeError(
            f"No latest 30-day bar found for UI-selected date {target_date_total.strftime(ISO_DATE_FORMAT)} for {location!r}."
        )
    logging.info(
        "30-day last bar | UI date=%s | website date=%s | Y-axis value=%s m3",
        target_date_total.strftime(ISO_DATE_FORMAT),
        target_date_total.strftime(ISO_DATE_FORMAT),
        daily,
    )

    # Min/max come from the 15-minute diagram at Selected_Date (zero-day offset).
    select_date_for_metric(page, "MIN/MAX DNEVNA POTROŠNJA (m3) / 15-minute", target_date_interval)
    select_interval(page, env("INTERVAL_1_DAY_LABEL", "1 Day (15-minutely)") or "1 Day (15-minutely)")
    payload_1_day = extract_chart_payload(page)
    values_1_day = numbers_from_chart_payload(payload_1_day)
    if not values_1_day:
        save_debug_artifacts(page, debug_dir, location, "1day_no_values")
        raise RuntimeError(f"No numeric 1-day chart values found for {location!r}.")
    min_daily = min(values_1_day)
    max_daily = max(values_1_day)

    logging.info(
        "Read %s: daily=%s m3, max=%s m3, min=%s m3",
        location,
        daily,
        max_daily,
        min_daily,
    )
    return Measurement(
        daily_m3=daily,
        max_daily_m3=max_daily,
        min_daily_m3=min_daily,
        battery_voltage=battery_voltage,
    )


def run_browser(
    jobs: list[StationJob],
    headless: bool,
    slow_mo_ms: int,
    keep_browser_open: bool,
    debug_dir: Path,
    limit: int | None,
    target_date_interval: datetime,
    target_date_total: datetime,
) -> RunResult:
    base_url = env("URL") or env("url")
    email = env("EMAIL") or env("GMAIL")
    password = env("PASSWORD")
    if not base_url or not email or not password:
        raise RuntimeError(".env must define url, email/gmail, and password.")

    selected_jobs = jobs[:limit] if limit else jobs
    measurements: dict[tuple[str, str], Measurement] = {}
    successes: list[StationJob] = []
    failures: list[RunIssue] = []
    no_data: list[RunIssue] = []

    with sync_playwright() as p:
        launch_kwargs: dict[str, Any] = {"headless": headless, "slow_mo": slow_mo_ms}
        chrome_path = env("CHROME_PATH")
        if chrome_path and Path(chrome_path).exists():
            launch_kwargs["executable_path"] = chrome_path

        logging.info("Launching browser: headless=%s slow_mo_ms=%s", headless, slow_mo_ms)
        browser: Browser = p.chromium.launch(**launch_kwargs)
        context = browser.new_context(viewport={"width": 1440, "height": 950})
        page = context.new_page()
        try:
            login(page, base_url, email, password)
            for idx, job in enumerate(selected_jobs, start=1):
                logging.info(
                    "[%s/%s] Station=%s, Excel row=%s, device=%r",
                    idx,
                    len(selected_jobs),
                    job.label,
                    job.excel_row,
                    job.station.uredjaj,
                )
                try:
                    measurement = scrape_measurement(
                        page, job.station.uredjaj, target_date_interval, target_date_total, debug_dir
                    )
                    measurements[job.key] = measurement
                    successes.append(job)
                    logging.info(
                        'FOUND: MIN: %s MAX: %s DAILY: %s BATTERY: %s for "%s"',
                        measurement.min_daily_m3,
                        measurement.max_daily_m3,
                        measurement.daily_m3,
                        measurement.battery_voltage,
                        job.label,
                    )
                except Exception:
                    reason = "Unknown error"
                    try:
                        reason = one_line(sys.exc_info()[1])
                    except Exception:
                        pass
                    # A single station failing is not fatal, so keep it at
                    # warning level here and let the end-of-run summary raise
                    # it to an error with the full explanation. The traceback
                    # stays available under --verbose.
                    logging.warning(
                        "Skipping station %s (Excel row %s): %s",
                        job.label,
                        job.excel_row,
                        one_line(reason, 200),
                    )
                    # --verbose is on by default in the launcher, so tying the
                    # traceback to it would bury the warning above under a wall
                    # of Python frames. The reason is kept in full for the
                    # end-of-run summary; set SHOW_TRACEBACKS=1 to get frames.
                    if as_bool(env("SHOW_TRACEBACKS"), default=False):
                        logging.warning(
                            "Traceback for station %s (Excel row %s)", job.label, job.excel_row, exc_info=True
                        )
                    issue = RunIssue(job, reason)
                    if "No numeric 1-day chart values" in reason or "No 30-day bar found" in reason:
                        no_data.append(issue)
                    else:
                        failures.append(issue)
                    if as_bool(env("CONTINUE_ON_ERROR"), default=True):
                        continue
                    raise
        finally:
            if keep_browser_open and not headless:
                logging.info("KEEP_BROWSER_OPEN is enabled. Press Ctrl+C in this terminal when done inspecting.")
                try:
                    while True:
                        time.sleep(1)
                except KeyboardInterrupt:
                    pass
            context.close()
            browser.close()

    return RunResult(measurements=measurements, successes=successes, failures=failures, no_data=no_data)


def merge_run_results(results: Iterable[RunResult]) -> RunResult:
    measurements: dict[tuple[str, str], Measurement] = {}
    successes: list[StationJob] = []
    failures: list[RunIssue] = []
    no_data: list[RunIssue] = []
    for result in results:
        measurements.update(result.measurements)
        successes.extend(result.successes)
        failures.extend(result.failures)
        no_data.extend(result.no_data)
    return RunResult(measurements=measurements, successes=successes, failures=failures, no_data=no_data)


def split_jobs(jobs: list[StationJob], workers: int) -> list[list[StationJob]]:
    chunks = [[] for _ in range(workers)]
    for idx, job in enumerate(jobs):
        chunks[idx % workers].append(job)
    return [chunk for chunk in chunks if chunk]


def run_browser_parallel(
    jobs: list[StationJob],
    workers: int,
    headless: bool,
    slow_mo_ms: int,
    keep_browser_open: bool,
    debug_dir: Path,
    limit: int | None,
    target_date_interval: datetime,
    target_date_total: datetime,
) -> RunResult:
    selected_jobs = jobs[:limit] if limit else jobs
    if workers <= 1 or len(selected_jobs) <= 1:
        return run_browser(
            jobs=selected_jobs,
            headless=headless,
            slow_mo_ms=slow_mo_ms,
            keep_browser_open=keep_browser_open,
            debug_dir=debug_dir,
            limit=None,
            target_date_interval=target_date_interval,
            target_date_total=target_date_total,
        )

    workers = min(workers, len(selected_jobs))
    if not headless:
        logging.warning("Running %s headed browsers in parallel. Use --headless for faster, less noisy parallel runs.", workers)
    if keep_browser_open:
        logging.warning("--keep-browser-open is ignored for parallel runs.")
        keep_browser_open = False

    chunks = split_jobs(selected_jobs, workers)
    logging.info("Running %s jobs across %s browser workers.", len(selected_jobs), len(chunks))
    results: list[RunResult] = []
    with ThreadPoolExecutor(max_workers=len(chunks)) as executor:
        future_by_chunk = {
            executor.submit(
                run_browser,
                jobs=chunk,
                headless=headless,
                slow_mo_ms=slow_mo_ms,
                keep_browser_open=keep_browser_open,
                debug_dir=debug_dir / f"worker-{idx + 1}",
                limit=None,
                target_date_interval=target_date_interval,
                target_date_total=target_date_total,
            ): chunk
            for idx, chunk in enumerate(chunks)
        }
        for future in as_completed(future_by_chunk):
            chunk = future_by_chunk[future]
            try:
                results.append(future.result())
            except Exception as exc:
                logging.exception("Browser worker failed for %s job(s).", len(chunk))
                failures = [RunIssue(job, str(exc)) for job in chunk]
                results.append(RunResult(measurements={}, successes=[], failures=failures, no_data=[]))

    return merge_run_results(results)


def log_run_report(result: RunResult) -> None:
    logging.info("")
    logging.info("========== EXECUTION REPORT ==========")
    logging.info("SUCCESSFUL: %s", len(result.successes))
    for job in result.successes:
        measurement = result.measurements.get(job.key)
        logging.info(
            "  OK | row=%s | %s | daily=%s max=%s min=%s battery=%s",
            job.excel_row,
            job.label,
            measurement.daily_m3 if measurement else None,
            measurement.max_daily_m3 if measurement else None,
            measurement.min_daily_m3 if measurement else None,
            measurement.battery_voltage if measurement else None,
        )

    # Counts stay at info when they are zero so a clean run reads calm; the
    # buckets that actually need attention are raised to warning/error, which
    # is what colours them orange/red in the launcher and the terminal.
    no_data_log = logging.warning if result.no_data else logging.info
    no_data_log("NO DATA / NO ENTRIES: %s", len(result.no_data))
    for issue in result.no_data:
        logging.warning(
            "  NO DATA | row=%s | %s | %s", issue.job.excel_row, issue.job.label, one_line(issue.reason)
        )

    failed_log = logging.error if result.failures else logging.info
    failed_log("FAILED: %s", len(result.failures))
    for issue in result.failures:
        logging.error(
            "  FAIL | row=%s | %s | %s", issue.job.excel_row, issue.job.label, one_line(issue.reason)
        )
    logging.info("======================================")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape EcoKing consumption values for a selected reporting date.")
    parser.add_argument("--output", "--workbook", dest="output", default=env("OUTPUT_PATH"), help="Standalone output workbook path. Defaults to the Desktop report.")
    parser.add_argument("--template", default=env("TEMPLATE_PATH", DEFAULT_TEMPLATE), help="Master blank template workbook path.")
    parser.add_argument("--stations", default=env("STATIONS_PATH") or env("LOCATION_MAP_PATH"), help="Station registry JSON. Defaults to stations.json beside the script.")
    parser.add_argument("--headless", action="store_true", help="Run browser hidden.")
    parser.add_argument("--headed", action="store_true", help="Run browser visible.")
    parser.add_argument("--slow-mo-ms", type=int, default=int(env("SLOW_MO_MS", "0") or "0"), help="Delay browser actions so clicks are visible.")
    parser.add_argument("--workers", type=int, default=int(env("WORKERS", "1") or "1"), help="Number of parallel browser workers. Default: 1.")
    parser.add_argument("--keep-browser-open", action="store_true", help="Keep headed browser open after the run.")
    parser.add_argument("--verbose", action="store_true", default=as_bool(env("VERBOSE"), default=True), help="Enable verbose logs.")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N stations.")
    parser.add_argument("--selected-date", default=None, help="Reporting date in YYYY-MM-DD format. Defaults to yesterday.")
    return parser.parse_args()


def main() -> int:
    load_dotenv()
    args = parse_args()
    configure_logging(args.verbose)

    headless_default = as_bool(env("HEADLESS"), default=False)
    headless = True if args.headless else False if args.headed else headless_default

    template_path = Path(args.template)
    if not template_path.exists():
        raise FileNotFoundError(template_path)
    target_date = parse_selected_date(args.selected_date)
    default_output = (desktop_directory() or Path.cwd()) / f"EcoKing_Report_{target_date.strftime(ISO_DATE_FORMAT)}.xlsx"
    output_path = Path(args.output) if args.output else default_output
    target_date_interval, target_date_total = target_dates(target_date)
    logging.info(
        "Selected date=%s; interval target=%s; total target=%s",
        target_date.strftime(WEBSITE_DATE_FORMAT),
        target_date_interval.strftime(WEBSITE_DATE_FORMAT),
        target_date_total.strftime(WEBSITE_DATE_FORMAT),
    )

    keep_browser_open = args.keep_browser_open or as_bool(env("KEEP_BROWSER_OPEN"), default=False)

    stations = load_stations(args.stations)
    rows = load_template_rows(template_path)
    logging.info("Loaded %s workbook rows from %s", len(rows), template_path)
    jobs = build_station_jobs(stations, rows)
    logging.info("Built %s station jobs from %s stations", len(jobs), len(stations))
    if not jobs:
        raise RuntimeError("No station matched a template row. Open the station list and fix the entries.")

    result = run_browser_parallel(
        jobs=jobs,
        workers=max(1, args.workers),
        headless=headless,
        slow_mo_ms=args.slow_mo_ms,
        keep_browser_open=keep_browser_open,
        debug_dir=Path("debug"),
        limit=args.limit,
        target_date_interval=target_date_interval,
        target_date_total=target_date_total,
    )

    report_path = clone_and_populate_template(
        template_path, output_path, target_date, result.measurements, jobs
    )
    if as_bool(env("COPY_REPORT_TO_DESKTOP"), default=True):
        report_path = store_report_on_desktop(report_path)
    log_run_report(result)
    logging.info(
        "REPORT GENERATED SUCCESSFULLY: %s for %s with %s mapped rows",
        report_path,
        target_date.strftime(ISO_DATE_FORMAT),
        len(result.measurements),
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit(130)
    except Exception as exc:
        logging.exception("Run failed: %s", exc)
        raise SystemExit(1)
