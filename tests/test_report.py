import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import ecoking_daily as daily
from ecoking.stations import ExcelRow, Station

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "ECO KING BLANKO TABLICA.xlsx"


class StationJobTests(unittest.TestCase):
    rows = [
        ExcelRow(row=22, lokacija="REZERVOAR BAJER 1", vodomjer="ULAZ"),
        ExcelRow(row=23, lokacija="REZERVOAR BAJER 1", vodomjer="IZLAZ"),
    ]

    def test_a_station_is_paired_with_its_row(self) -> None:
        jobs = daily.build_station_jobs([Station("REZERVOAR BAJER 1", "IZLAZ", "Bajer 1 - I")], self.rows)
        self.assertEqual([job.excel_row for job in jobs], [23])

    def test_a_disabled_station_is_skipped(self) -> None:
        stations = [Station("REZERVOAR BAJER 1", "ULAZ", "Bajer 1 - U", enabled=False)]
        self.assertEqual(daily.build_station_jobs(stations, self.rows), [])

    def test_a_station_without_a_device_name_is_skipped(self) -> None:
        self.assertEqual(daily.build_station_jobs([Station("REZERVOAR BAJER 1", "ULAZ", "")], self.rows), [])

    def test_a_station_with_no_matching_row_is_skipped(self) -> None:
        self.assertEqual(daily.build_station_jobs([Station("NEMA", "ULAZ", "X")], self.rows), [])

    def test_matching_ignores_case_and_accents(self) -> None:
        jobs = daily.build_station_jobs([Station("rezervoar bajer 1", "ulaz", "Bajer 1 - U")], self.rows)
        self.assertEqual([job.excel_row for job in jobs], [22])


class MeasurementTests(unittest.TestCase):
    def test_litres_per_second_conversions(self) -> None:
        measurement = daily.Measurement(daily_m3=86.4, max_daily_m3=0.9, min_daily_m3=0.09)
        self.assertAlmostEqual(measurement.daily_lps, 1.0)
        self.assertAlmostEqual(measurement.max_daily_lps, 1.0)
        self.assertAlmostEqual(measurement.min_daily_lps, 0.1)

    def test_missing_values_stay_missing(self) -> None:
        measurement = daily.Measurement(daily_m3=None, max_daily_m3=None, min_daily_m3=None)
        self.assertIsNone(measurement.daily_lps)
        self.assertIsNone(measurement.max_daily_lps)
        self.assertIsNone(measurement.min_daily_lps)


@unittest.skipUnless(TEMPLATE.exists(), "The report template is not available.")
class ReportWritingTests(unittest.TestCase):
    """The report is a copy of the template with four cells filled per station."""

    def build(self, measurements, stations):
        rows = daily.load_template_rows(TEMPLATE)
        jobs = daily.build_station_jobs(stations, rows)
        keyed = {job.key: measurements[index] for index, job in enumerate(jobs)}
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        output = Path(directory.name) / "report.xlsx"
        daily.clone_and_populate_template(TEMPLATE, output, datetime(2026, 7, 23), keyed, jobs)
        return load_workbook(output, data_only=False), jobs

    def test_values_land_in_the_columns_the_template_expects(self) -> None:
        measurement = daily.Measurement(daily_m3=120.5, max_daily_m3=3.25, min_daily_m3=0.5, battery_voltage="3.6V")
        workbook, jobs = self.build([measurement], [Station("REZERVOAR BAJER 1", "ULAZ", "Bajer 1 - U")])
        sheet = workbook.active
        row = jobs[0].excel_row

        self.assertEqual(sheet.cell(row=row, column=daily.COLUMN_DAILY_M3).value, 120.5)
        self.assertEqual(sheet.cell(row=row, column=daily.COLUMN_MAX_DAILY_M3).value, 3.25)
        self.assertEqual(sheet.cell(row=row, column=daily.COLUMN_MIN_DAILY_M3).value, 0.5)
        self.assertEqual(sheet.cell(row=row, column=daily.COLUMN_BATTERY).value, "3.6V")

    def test_the_litres_per_second_formulas_survive(self) -> None:
        measurement = daily.Measurement(daily_m3=10.0, max_daily_m3=1.0, min_daily_m3=0.1)
        workbook, jobs = self.build([measurement], [Station("REZERVOAR BAJER 1", "ULAZ", "Bajer 1 - U")])
        sheet = workbook.active
        row = jobs[0].excel_row
        self.assertEqual(sheet.cell(row=row, column=7).value, f"=F{row}*1000/86400")
        self.assertEqual(sheet.cell(row=row, column=9).value, f"=H{row}*1000/900")

    def test_the_sheet_is_renamed_and_stamped_with_the_date(self) -> None:
        measurement = daily.Measurement(daily_m3=1.0, max_daily_m3=1.0, min_daily_m3=1.0)
        workbook, _ = self.build([measurement], [Station("REZERVOAR BAJER 1", "ULAZ", "Bajer 1 - U")])
        self.assertEqual(workbook.active.title, "2026-07-23")
        self.assertEqual(workbook.active["N1"].value, "DATUM: 2026-07-23")

    def test_rows_without_a_measurement_are_left_empty(self) -> None:
        measurement = daily.Measurement(daily_m3=7.0, max_daily_m3=1.0, min_daily_m3=1.0)
        workbook, jobs = self.build([measurement], [Station("REZERVOAR BAJER 1", "ULAZ", "Bajer 1 - U")])
        sheet = workbook.active
        other = jobs[0].excel_row + 1
        self.assertIsNone(sheet.cell(row=other, column=daily.COLUMN_DAILY_M3).value)

    def test_writing_over_the_template_is_refused(self) -> None:
        with self.assertRaises(RuntimeError):
            daily.clone_and_populate_template(TEMPLATE, TEMPLATE, datetime(2026, 7, 23), {}, [])


if __name__ == "__main__":
    unittest.main()
