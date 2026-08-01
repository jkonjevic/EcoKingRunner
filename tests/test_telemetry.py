import argparse
import os
import shutil
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import ecoking_daily as daily
from ecoking import telemetry

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "ECO KING BLANKO TABLICA.xlsx"

TARGET = datetime(2026, 7, 31)

#: Two chambers on one page -- each meter has its own level column.
KUMBOR_TABLE = {
    "headers": [
        "Br.",
        "Vrijeme",
        "M1 Kolicina (m3)",
        "M1 Protok (l/s)",
        "M1 Nivo (m)",
        "M2 Kolicina (m3)",
        "M2 Protok (l/s)",
        "M2 Nivo (m)",
        "Napon (V)",
    ],
    "rows": [
        ["6", "2026-07-31 18:00:01", "1513967.96", "0.00", "3.70", "1533508.35", "0.00", "4.20", "24.2"],
        ["7", "2026-07-31 17:00:01", "1513967.96", "0.00", "3.63", "1533508.35", "0.00", "4.11", "24.2"],
        ["8", "2026-07-31 16:00:01", "1513967.96", "0.00", "3.55", "1533508.35", "0.00", "4.02", "24.2"],
    ],
}

SINGLE_TABLE = {
    "headers": ["Br.", "Vrijeme", "Kolicina (m3)", "Protok (l/s)", "Nivo (m)", "Napon (V)"],
    "rows": [
        ["1", "2026-07-31 18:00:01", "1143054.78", "0.00", "1.80", "24.3"],
        ["2", "2026-07-31 17:00:01", "1143054.78", "0.00", "1.73", "24.3"],
        ["3", "2026-07-30 17:00:01", "1143000.00", "0.00", "9.99", "24.3"],
    ],
}


class MatchingTests(unittest.TestCase):
    def test_match_key_strips_the_antenna_alt_text(self):
        # The icon's alt runs into the name with and without a space.
        self.assertEqual(telemetry.match_key("Rezervoar Podi Tele"), "rezervoar podi")
        self.assertEqual(telemetry.match_key("Rezervoar KulaTele"), "rezervoar kula")
        self.assertEqual(telemetry.match_key("Rezervoar Kula"), "rezervoar kula")

    def test_match_key_ignores_diacritics(self):
        self.assertEqual(
            telemetry.match_key("Rezervoar Spanjola mjerač Tele"),
            telemetry.match_key("Rezervoar Spanjola mjerac"),
        )

    def test_meter_number_survives_the_glued_alt_text(self):
        self.assertEqual(telemetry.meter_number("Rezervoar Kumbor mjerač 2Tele"), 2)
        self.assertEqual(telemetry.meter_number("Rezervoar Kumbor mjerač 1Tele"), 1)
        self.assertIsNone(telemetry.meter_number("Rezervoar Podi Tele"))


class ResolveLocationTests(unittest.TestCase):
    mapping = {
        "Rezervoar Podi Tele": "REZERVOAR PODI",
        "Rezervoar Bajer2 mjerač 2 Tele": "REZERVOAR BAJER 2",
        "Rezervoar Kumbor mjerač 1Tele": "REZERVOAR KUMBOR LIJEVA KOMORA",
        "Rezervoar Kumbor mjerač 2Tele": "REZERVOAR KUMBOR DESNA KOMORA",
    }

    def test_exact_match(self):
        self.assertEqual(
            telemetry.resolve_workbook_location("Rezervoar Podi Tele", self.mapping),
            "REZERVOAR PODI",
        )

    def test_other_meter_of_a_single_level_page_resolves(self):
        # Bajer2's two meters share one level, so either name is the same row.
        self.assertEqual(
            telemetry.resolve_workbook_location("Rezervoar Bajer2 mjerač 1 Tele", self.mapping),
            "REZERVOAR BAJER 2",
        )

    def test_separately_mapped_meters_stay_separate(self):
        self.assertEqual(
            telemetry.resolve_workbook_location("Rezervoar Kumbor mjerač 2Tele", self.mapping),
            "REZERVOAR KUMBOR DESNA KOMORA",
        )

    def test_unmapped_location_is_reported(self):
        self.assertIsNone(telemetry.resolve_workbook_location("Rezervoar Zmijica Tele", self.mapping))


class FindLevelTests(unittest.TestCase):
    def test_picks_the_meters_own_level_column(self):
        self.assertEqual(telemetry.find_level(KUMBOR_TABLE, TARGET, 1)[0], 3.63)
        self.assertEqual(telemetry.find_level(KUMBOR_TABLE, TARGET, 2)[0], 4.11)

    def test_single_level_column_ignores_the_meter_number(self):
        self.assertEqual(telemetry.find_level(SINGLE_TABLE, TARGET, 1)[0], 1.73)
        self.assertEqual(telemetry.find_level(SINGLE_TABLE, TARGET, None)[0], 1.73)

    def test_only_the_selected_date_counts(self):
        value, stamp = telemetry.find_level(SINGLE_TABLE, TARGET, None)
        self.assertEqual(stamp, "2026-07-31 17:00:01")
        self.assertNotEqual(value, 9.99)

    def test_missing_hour_is_an_error(self):
        with self.assertRaises(RuntimeError):
            telemetry.find_level(SINGLE_TABLE, datetime(2026, 7, 29), None)


class ConfigTests(unittest.TestCase):
    def test_locations_and_mapping_load(self):
        locations = telemetry.load_locations(ROOT)
        mapping = telemetry.load_mapping(ROOT)
        self.assertIn("Rezervoar Podi Tele", locations)
        self.assertTrue(mapping)

    def test_every_mapped_location_exists_in_the_template(self):
        workbook = load_workbook(TEMPLATE, data_only=True, read_only=True)
        try:
            labels = {
                telemetry.normalize(cell.value)
                for row in workbook.worksheets[0].iter_rows(min_row=2, min_col=2, max_col=2)
                for cell in row
                if cell.value
            }
        finally:
            workbook.close()
        for workbook_location in telemetry.load_mapping(ROOT).values():
            self.assertIn(telemetry.normalize(workbook_location), labels, workbook_location)


class TelemetryHeadlessTests(unittest.TestCase):
    """The telemetry browser is a separate decision from the EcoKing one."""

    def setUp(self):
        self.previous = os.environ.pop("TELEMETRY_HEADLESS", None)
        self.addCleanup(self._restore)

    def _restore(self):
        os.environ.pop("TELEMETRY_HEADLESS", None)
        if self.previous is not None:
            os.environ["TELEMETRY_HEADLESS"] = self.previous

    @staticmethod
    def args(**overrides):
        defaults = {"telemetry_headed": False, "telemetry_headless": False}
        return argparse.Namespace(**{**defaults, **overrides})

    def test_follows_the_ecoking_pass_when_nothing_is_configured(self):
        self.assertTrue(daily.telemetry_headless(self.args(), headless=True))
        self.assertFalse(daily.telemetry_headless(self.args(), headless=False))

    def test_env_overrides_the_ecoking_pass(self):
        os.environ["TELEMETRY_HEADLESS"] = "0"
        self.assertFalse(daily.telemetry_headless(self.args(), headless=True))
        os.environ["TELEMETRY_HEADLESS"] = "1"
        self.assertTrue(daily.telemetry_headless(self.args(), headless=False))

    def test_explicit_flags_win_over_env(self):
        os.environ["TELEMETRY_HEADLESS"] = "1"
        self.assertFalse(daily.telemetry_headless(self.args(telemetry_headed=True), headless=True))
        os.environ["TELEMETRY_HEADLESS"] = "0"
        self.assertTrue(daily.telemetry_headless(self.args(telemetry_headless=True), headless=False))


class WriteLevelsTests(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.report = self.tmp / "report.xlsx"
        shutil.copy2(TEMPLATE, self.report)
        self.addCleanup(shutil.rmtree, self.tmp, True)

    def test_level_lands_on_every_meter_row_of_the_location(self):
        readings = {
            "Rezervoar Banja1 mjerač Tele": telemetry.LevelReading(
                "Rezervoar Banja1 mjerač Tele", 3.66, "2026-07-31 17:00:01"
            )
        }
        result = telemetry.write_levels(
            self.report, readings, {"Rezervoar Banja1 mjerač Tele": "REZERVOAR BANJSKI 1"}
        )
        worksheet = load_workbook(self.report).worksheets[0]
        # Rows 7/8 are the ULAZ and IZLAZ meters of REZERVOAR BANJSKI 1.
        self.assertEqual(worksheet.cell(7, telemetry.LEVEL_COLUMN_INDEX).value, 3.66)
        self.assertEqual(worksheet.cell(8, telemetry.LEVEL_COLUMN_INDEX).value, 3.66)
        self.assertEqual(result.written_rows, 2)

    def test_formulas_and_other_columns_are_untouched(self):
        readings = {
            "Rezervoar Podi Tele": telemetry.LevelReading(
                "Rezervoar Podi Tele", 1.73, "2026-07-31 17:00:01"
            )
        }
        telemetry.write_levels(self.report, readings, {"Rezervoar Podi Tele": "REZERVOAR PODI"})
        worksheet = load_workbook(self.report).worksheets[0]
        self.assertEqual(worksheet.cell(28, 7).value, "=F28*1000/86400")
        self.assertEqual(worksheet.cell(28, 12).value, 1.73)

    def test_unmapped_reading_is_recorded_not_written(self):
        readings = {
            "Rezervoar Zmijica Tele": telemetry.LevelReading(
                "Rezervoar Zmijica Tele", 2.94, "2026-07-31 17:00:01"
            )
        }
        result = telemetry.write_levels(self.report, readings, {})
        self.assertEqual(result.written_rows, 0)
        self.assertEqual(len(result.failures), 1)


if __name__ == "__main__":
    unittest.main()
