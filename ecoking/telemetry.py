"""Second scraping pass: reservoir level at 17:00 from the ViK telemetry site.

The EcoKing run fills the consumption columns, but the reservoir levels live on
a different system (``nadzorhnvik``), so they are collected once the workbook
already exists and written into its ``NIVO REZERVOARA U 17h`` column.

Two config files next to the app drive this:

* ``telemetry_list.py`` -- ``locations``: the rows of the site's "Spisak
  mjernih mjesta" table to visit, spelled exactly as they read on screen. The
  trailing ``Tele`` comes from the antenna icon's ``alt`` text, so it is
  stripped during matching rather than being required.
* ``telemetry_mapping.json`` -- site ``Lokacija`` -> workbook ``LOKACIJA``.

A level describes the reservoir, not one meter, so every meter row of a mapped
workbook location (``ULAZ`` and ``IZLAZ`` alike) gets the same value.
"""

from __future__ import annotations

import ast
import json
import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

LIST_FILE = "telemetry_list.py"
MAPPING_FILE = "telemetry_mapping.json"

LEVEL_HOUR = 17
LEVEL_COLUMN_HEADER = "NIVO REZERVOARA U 17h"
#: Fallback only; the header above is what is actually looked up.
LEVEL_COLUMN_INDEX = 12

DEFAULT_WAIT_MS = 10_000
NAVIGATION_TIMEOUT_MS = 60_000

_TABLE_CONTAINER = "#MtableID"
_METER_RE = re.compile(r"\bmjerac\s*(\d+)\b")
_TRAILING_TELE_RE = re.compile(r"\s*tele$")
_TIME_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M")

#: Pulls the whole data table out of the iframe in one round trip. The page
#: splits the header row and the data rows into two sibling tables, so headers
#: and cells are collected separately and lined up by position.
_READ_TABLE_JS = """
() => {
  const headers = [...document.querySelectorAll('th')].map((th) => th.textContent.trim());
  const rows = [...document.querySelectorAll('tr')]
    .map((tr) => [...tr.querySelectorAll('td')].map((td) => td.textContent.trim()))
    .filter((cells) => cells.length > 1);
  return { headers, rows };
}
"""


@dataclass(frozen=True)
class LevelReading:
    """One ``Nivo`` value read off a location's detail table."""

    location: str
    value: float
    timestamp: str


@dataclass
class TelemetryResult:
    readings: dict[str, LevelReading] = field(default_factory=dict)
    failures: list[tuple[str, str]] = field(default_factory=list)
    written_rows: int = 0
    written_locations: list[str] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Text matching
# --------------------------------------------------------------------------- #


def normalize(text: object) -> str:
    """Casefold, strip diacritics and collapse whitespace for comparisons."""
    decomposed = unicodedata.normalize("NFKD", str(text))
    stripped = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", stripped).strip().lower()


def match_key(text: object) -> str:
    """Normalized form with the antenna icon's ``Tele`` suffix removed.

    ``telemetry_list.py`` was written from the rendered page, where the icon's
    ``alt`` text runs into the location name (``"Rezervoar KulaTele"``), while
    the DOM text is just ``"Rezervoar Kula"``. Both sides go through here.
    """
    return _TRAILING_TELE_RE.sub("", normalize(text)).strip()


def meter_number(location: str) -> int | None:
    """``"Rezervoar Kumbor mjerač 2"`` -> ``2``; ``None`` when unnumbered.

    Matched against :func:`match_key`, not the raw name: the icon's ``Tele``
    runs straight into the digit (``"mjerač 2Tele"``) and would otherwise hide
    the meter number -- which is what picks ``M2 Nivo`` over ``M1 Nivo`` on the
    pages that report two chambers.
    """
    found = _METER_RE.search(match_key(location))
    return int(found.group(1)) if found else None


def _without_meter(location: str) -> str:
    return _METER_RE.sub("", match_key(location)).strip()


# --------------------------------------------------------------------------- #
# Config files
# --------------------------------------------------------------------------- #


def load_locations(root: Path) -> list[str]:
    """Read ``locations`` out of ``telemetry_list.py`` without importing it."""
    path = root / LIST_FILE
    if not path.exists():
        raise FileNotFoundError(f"{LIST_FILE} not found next to the app ({path}).")
    tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
    for node in tree.body:
        targets = node.targets if isinstance(node, ast.Assign) else []
        if any(isinstance(target, ast.Name) and target.id == "locations" for target in targets):
            values = ast.literal_eval(node.value)
            return [str(value).strip() for value in values if str(value).strip()]
    raise ValueError(f"{LIST_FILE} does not define a `locations` list.")


def load_mapping(root: Path) -> dict[str, str]:
    """Read the site ``Lokacija`` -> workbook ``LOKACIJA`` mapping."""
    path = root / MAPPING_FILE
    if not path.exists():
        raise FileNotFoundError(f"{MAPPING_FILE} not found next to the app ({path}).")
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict):
        raise ValueError(f"{MAPPING_FILE} must be an object of site name -> workbook LOKACIJA.")
    return {str(key): str(value) for key, value in raw.items()}


def resolve_workbook_location(location: str, mapping: dict[str, str]) -> str | None:
    """Find the workbook ``LOKACIJA`` a site location belongs to.

    An exact (normalized) key wins. Failing that, the ``mjerač N`` suffix is
    dropped and the reduced name is used -- but only when it picks out exactly
    one mapping entry. That resolves ``Bajer2 mjerač 1`` against a mapping
    written for ``Bajer2 mjerač 2`` (one shared level, either meter reads it)
    while leaving Kumbor's two separately mapped chambers alone.
    """
    by_key = {match_key(key): value for key, value in mapping.items()}
    exact = by_key.get(match_key(location))
    if exact:
        return exact

    reduced = _without_meter(location)
    candidates = {value for key, value in by_key.items() if _without_meter(key) == reduced}
    if len(candidates) == 1:
        return candidates.pop()
    return None


# --------------------------------------------------------------------------- #
# Reading the detail table
# --------------------------------------------------------------------------- #


def parse_number(text: str) -> float | None:
    cleaned = str(text).strip().replace(",", ".")
    if not cleaned or set(cleaned) <= {"-", "."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_timestamp(text: str) -> datetime | None:
    cleaned = re.sub(r"\s+", " ", str(text)).strip()
    for time_format in _TIME_FORMATS:
        try:
            return datetime.strptime(cleaned, time_format)
        except ValueError:
            continue
    return None


def level_column_index(headers: list[str], meter: int | None) -> int | None:
    """Pick the ``Nivo`` column, honouring ``M1``/``M2`` when the page has both."""
    candidates = [index for index, header in enumerate(headers) if "nivo" in normalize(header)]
    if not candidates:
        return None
    if len(candidates) > 1 and meter is not None:
        for index in candidates:
            if re.search(rf"\bm\s*{meter}\b", normalize(headers[index])):
                return index
    return candidates[0]


def find_level(table: dict[str, Any], target_date: datetime, meter: int | None) -> tuple[float, str]:
    """Return the ``Nivo`` value (and its timestamp) at 17:00 on ``target_date``."""
    headers: list[str] = [str(value) for value in table.get("headers") or []]
    rows: list[list[str]] = [[str(cell) for cell in row] for row in table.get("rows") or []]
    if not headers or not rows:
        # These reasons reach the operator through the run log, which is in
        # Serbian, so they are written in it rather than translated later.
        raise RuntimeError("Tabela sa podacima je prazna.")

    column = level_column_index(headers, meter)
    if column is None:
        raise RuntimeError(f"Nema kolone 'Nivo' u tabeli (kolone: {', '.join(headers)}).")
    time_column = next(
        (index for index, header in enumerate(headers) if "vrijeme" in normalize(header)), 1
    )

    wanted = target_date.date()
    matches: list[tuple[datetime, list[str]]] = []
    for cells in rows:
        if time_column >= len(cells):
            continue
        stamp = parse_timestamp(cells[time_column])
        if stamp and stamp.date() == wanted and stamp.hour == LEVEL_HOUR:
            matches.append((stamp, cells))
    if not matches:
        raise RuntimeError(f"Nema reda za {LEVEL_HOUR}:00 na dan {wanted.isoformat()} u tabeli.")

    # Readings land a second or two past the hour, so take the earliest 17:xx
    # row -- that is the one the operators read as "the 17h value".
    stamp, cells = min(matches, key=lambda item: (item[0].minute, item[0].second))
    if column >= len(cells):
        raise RuntimeError(f"Red {stamp} nema ćeliju u koloni 'Nivo'.")
    value = parse_number(cells[column])
    if value is None:
        raise RuntimeError(f"'Nivo' je prazan u {stamp} (vrijednost: {cells[column]!r}).")
    return value, stamp.strftime("%Y-%m-%d %H:%M:%S")


# --------------------------------------------------------------------------- #
# Browser flow
# --------------------------------------------------------------------------- #


def login(page: Any, base_url: str, username: str, password: str) -> None:
    """Open the site and post the login form if it asks for credentials."""
    logging.info("Telemetry: opening %s", base_url)
    page.goto(base_url, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    if page.locator("input[name='username']").count() == 0:
        logging.info("Telemetry: already signed in")
        return
    page.fill("input[name='username']", username)
    page.fill("input[name='password']", password)
    page.click("input[type='submit']")
    page.wait_for_load_state("domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    if page.locator("input[name='username']").count() > 0:
        raise RuntimeError(
            "Prijava na telemetriju je odbijena; provjerite TELEMETRIJA_USERNAME/PASSWORD."
        )
    logging.info("Telemetry: signed in as %s", username)


def index_links(page: Any) -> list[str]:
    """Location link labels from "Spisak mjernih mjesta", in table order."""
    return [str(text) for text in page.locator(f"{_TABLE_CONTAINER} a").all_text_contents()]


def open_location(page: Any, link_index: int) -> None:
    """Click a location row and let its detail page settle."""
    page.locator(f"{_TABLE_CONTAINER} a").nth(link_index).click()
    page.wait_for_load_state("domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    page.wait_for_selector("#DataForm", timeout=NAVIGATION_TIMEOUT_MS)


def apply_date_filter(page: Any, target_date: datetime, wait_ms: int) -> dict[str, Any]:
    """Narrow the filter to ``target_date``, submit it and read the table back.

    The table lives in an iframe the filter form posts into, so the values only
    appear after the submit round trip -- which is slow enough that the wait is
    generous, then polled rather than assumed.
    """
    iso = target_date.strftime("%Y-%m-%d")
    page.fill("#dani_od", iso)
    page.fill("#dani_do", iso)
    page.click("#DataForm input[value='Prikaz']")

    deadline_ms = max(wait_ms, DEFAULT_WAIT_MS) * 3
    page.wait_for_timeout(wait_ms)
    waited = wait_ms
    table: dict[str, Any] = {}
    while True:
        frame = page.frame(name="tabelaIframe")
        if frame is not None:
            try:
                table = frame.evaluate(_READ_TABLE_JS)
            except Exception:
                table = {}
            if table.get("rows"):
                return table
        if waited >= deadline_ms:
            return table
        page.wait_for_timeout(1_000)
        waited += 1_000


def read_location_level(
    page: Any,
    index_url: str,
    link_index: int,
    location: str,
    target_date: datetime,
    wait_ms: int,
) -> LevelReading:
    open_location(page, link_index)
    table = apply_date_filter(page, target_date, wait_ms)
    value, stamp = find_level(table, target_date, meter_number(location))
    page.goto(index_url, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
    return LevelReading(location=location, value=value, timestamp=stamp)


def collect_levels(
    page: Any,
    locations: Iterable[str],
    target_date: datetime,
    index_url: str,
    wait_ms: int,
) -> TelemetryResult:
    """Visit every configured location and read its 17:00 level."""
    result = TelemetryResult()
    wanted = list(locations)
    labels = index_links(page)
    by_key: dict[str, int] = {}
    for index, label in enumerate(labels):
        by_key.setdefault(match_key(label), index)

    for position, location in enumerate(wanted, start=1):
        link_index = by_key.get(match_key(location))
        if link_index is None:
            logging.warning("Telemetry: %r is not in the site's location table; skipping.", location)
            result.failures.append((location, "Nema te lokacije u tabeli na sajtu."))
            continue
        logging.info("Telemetry [%s/%s] Location=%s", position, len(wanted), location)
        try:
            reading = read_location_level(
                page, index_url, link_index, location, target_date, wait_ms
            )
        except Exception as exc:
            reason = re.sub(r"\s+", " ", str(exc)).strip()[:200]
            logging.warning("Telemetry: could not read %s: %s", location, reason)
            result.failures.append((location, reason))
            try:
                page.goto(index_url, wait_until="domcontentloaded", timeout=NAVIGATION_TIMEOUT_MS)
            except Exception:
                pass
            continue
        result.readings[location] = reading
        logging.info("Telemetry: LEVEL %s m at %s for %s", reading.value, reading.timestamp, location)
    return result


def scrape_levels(
    target_date: datetime,
    base_url: str,
    username: str,
    password: str,
    locations: Iterable[str],
    headless: bool = True,
    slow_mo_ms: int = 0,
    wait_ms: int = DEFAULT_WAIT_MS,
) -> TelemetryResult:
    """Run the whole telemetry pass in its own browser."""
    from playwright.sync_api import sync_playwright

    with sync_playwright() as playwright:
        launch_kwargs: dict[str, Any] = {"headless": headless, "slow_mo": slow_mo_ms}
        chrome_path = os.getenv("CHROME_PATH")
        if chrome_path and Path(chrome_path).exists():
            launch_kwargs["executable_path"] = chrome_path
        browser = playwright.chromium.launch(**launch_kwargs)
        context = browser.new_context(viewport={"width": 1440, "height": 950})
        page = context.new_page()
        try:
            login(page, base_url, username, password)
            index_url = page.url
            return collect_levels(page, locations, target_date, index_url, wait_ms)
        finally:
            context.close()
            browser.close()


# --------------------------------------------------------------------------- #
# Writing the workbook
# --------------------------------------------------------------------------- #


def _location_rows(worksheet: Any, location_column: int) -> dict[str, list[int]]:
    """Map each normalized ``LOKACIJA`` to every row of its meter group.

    ``LOKACIJA`` is only written on the first row of a group (the cells below
    it are merged/blank), so it carries down until the next non-empty one.
    """
    groups: dict[str, list[int]] = {}
    current = ""
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=location_column).value
        label = str(cell).strip() if cell is not None else ""
        if label:
            current = label
        if not current:
            continue
        groups.setdefault(normalize(current), []).append(row)
    return groups


def write_levels(workbook_path: Path, readings: dict[str, LevelReading], mapping: dict[str, str]) -> TelemetryResult:
    """Write each reading into ``NIVO REZERVOARA U 17h`` for its workbook rows."""
    from openpyxl import load_workbook

    result = TelemetryResult(readings=readings)
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        worksheet = next(
            (
                candidate
                for candidate in workbook.worksheets
                if {str(cell.value).strip().upper() for cell in candidate[1] if cell.value}
                >= {"LOKACIJA", "VODOMJER"}
            ),
            None,
        )
        if worksheet is None:
            raise RuntimeError("Report is missing LOKACIJA and VODOMJER headers.")

        headers = {
            normalize(cell.value): cell.column for cell in worksheet[1] if cell.value is not None
        }
        location_column = headers.get(normalize("LOKACIJA"), 2)
        level_column = headers.get(normalize(LEVEL_COLUMN_HEADER), LEVEL_COLUMN_INDEX)
        groups = _location_rows(worksheet, location_column)

        for location, reading in readings.items():
            workbook_location = resolve_workbook_location(location, mapping)
            if not workbook_location:
                logging.warning(
                    "Telemetry: %r has no entry in %s; its level is not written.",
                    location,
                    MAPPING_FILE,
                )
                result.failures.append((location, f"Nema unosa u {MAPPING_FILE}."))
                continue
            rows = groups.get(normalize(workbook_location))
            if not rows:
                logging.warning(
                    "Telemetry: no report row with LOKACIJA=%r for %r.", workbook_location, location
                )
                result.failures.append((location, f"Nema reda LOKACIJA={workbook_location}."))
                continue
            for row in rows:
                worksheet.cell(row=row, column=level_column).value = reading.value
            result.written_rows += len(rows)
            result.written_locations.append(workbook_location)
            logging.info(
                "Telemetry: wrote %s m into %s (rows %s)",
                reading.value,
                workbook_location,
                ", ".join(str(row) for row in rows),
            )
        workbook.save(workbook_path)
    finally:
        workbook.close()
    return result


# --------------------------------------------------------------------------- #
# Stage entry point
# --------------------------------------------------------------------------- #


def _as_bool(value: str | None, default: bool) -> bool:
    if value is None or not value.strip():
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def run_stage(
    workbook_path: Path,
    target_date: datetime,
    root: Path,
    headless: bool | None = None,
    slow_mo_ms: int = 0,
    wait_ms: int | None = None,
) -> TelemetryResult | None:
    """Scrape the levels and write them into an already generated report.

    Never raises: the consumption report is finished by the time this runs, so
    a telemetry problem is logged and the report is kept as-is.
    """
    if not _as_bool(os.getenv("TELEMETRY_ENABLED"), True):
        logging.info("Telemetry stage is disabled (TELEMETRY_ENABLED=0); skipping.")
        return None

    base_url = os.getenv("TELEMETRIJA_URL")
    username = os.getenv("TELEMETRIJA_USERNAME")
    password = os.getenv("TELEMETRIJA_PASSWORD")
    if not base_url or not username or not password:
        logging.warning(
            "Telemetry stage skipped: .env must define TELEMETRIJA_URL, "
            "TELEMETRIJA_USERNAME and TELEMETRIJA_PASSWORD."
        )
        return None

    try:
        locations = load_locations(root)
        mapping = load_mapping(root)
    except Exception as exc:
        logging.warning("Telemetry stage skipped: %s", exc)
        return None

    if wait_ms is None:
        wait_ms = int(os.getenv("TELEMETRY_WAIT_MS") or DEFAULT_WAIT_MS)
    if headless is None:
        headless = _as_bool(os.getenv("TELEMETRY_HEADLESS"), True)

    logging.info(
        "Telemetry: reading %s:00 levels for %s across %s location(s)",
        LEVEL_HOUR,
        target_date.strftime("%Y-%m-%d"),
        len(locations),
    )
    try:
        scraped = scrape_levels(
            target_date=target_date,
            base_url=base_url,
            username=username,
            password=password,
            locations=locations,
            headless=bool(headless),
            slow_mo_ms=slow_mo_ms,
            wait_ms=wait_ms,
        )
    except Exception as exc:
        logging.warning(
            "Telemetry stage failed: %s. The report is kept without the 17h levels.",
            re.sub(r"\s+", " ", str(exc)).strip()[:200],
        )
        return None

    result = TelemetryResult(
        readings=scraped.readings,
        failures=list(scraped.failures),
    )
    if scraped.readings:
        try:
            written = write_levels(workbook_path, scraped.readings, mapping)
        except Exception as exc:
            logging.warning(
                "Telemetry: could not write the levels into %s: %s",
                workbook_path.name,
                re.sub(r"\s+", " ", str(exc)).strip()[:200],
            )
            return result
        result.written_rows = written.written_rows
        result.written_locations = written.written_locations
        result.failures.extend(written.failures)

    logging.info(
        "TELEMETRY DONE: %s of %s location(s) read, %s report row(s) filled",
        len(result.readings),
        len(locations),
        result.written_rows,
    )
    for location, reason in result.failures:
        logging.warning("TELEMETRY FAIL | %s | %s", location, reason)
    return result


def main() -> int:
    """Run the telemetry pass on its own, against an existing report."""
    import argparse

    from dotenv import load_dotenv

    root = Path(__file__).resolve().parent.parent
    load_dotenv(root / ".env")

    parser = argparse.ArgumentParser(description="Read 17h reservoir levels into an EcoKing report.")
    parser.add_argument("--workbook", required=True, help="Report .xlsx to fill in.")
    parser.add_argument("--date", required=True, help="Reporting date, YYYY-MM-DD.")
    parser.add_argument("--headed", action="store_true", help="Show the browser.")
    parser.add_argument("--wait-ms", type=int, default=None, help="Per-page settle time.")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")
    if args.wait_ms is not None:
        os.environ["TELEMETRY_WAIT_MS"] = str(args.wait_ms)

    result = run_stage(
        workbook_path=Path(args.workbook),
        target_date=datetime.strptime(args.date, "%Y-%m-%d"),
        root=root,
        headless=not args.headed,
    )
    return 0 if result and result.written_rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
