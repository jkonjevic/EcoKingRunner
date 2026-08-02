"""Station registry: the link between Excel template rows and EcoKing devices.

Each station says three things:

* ``lokacija`` / ``vodomjer`` -- the exact ``LOKACIJA`` and ``VODOMJER`` cells of
  the row in ``ECO KING BLANKO TABLICA.xlsx`` that the values are written into.
* ``uredjaj`` -- the device label as it reads on the EcoKing site, with the
  serial number and the ``Herceg Novi -`` prefix removed. ``Bajer 1 - U`` rather
  than ``358004092234384 - Herceg Novi - (R-BA) Bajer 1 - U``.

Because ``(LOKACIJA, VODOMJER)`` is unique in the template, the Excel lookup is
an exact dictionary hit rather than a similarity score.
"""

from __future__ import annotations

import json
import re
import unicodedata
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Iterable

SCHEMA_VERSION = 2

DEFAULT_STATIONS_FILE = "stations.json"
LEGACY_STATIONS_FILE = "herceg_novi_stations.json"

#: Site labels are prefixed with the city; it is the same for every device, so
#: it carries no information and is stripped from stored labels.
DEFAULT_CITY_PREFIXES = ("Herceg Novi",)

_SERIAL_PREFIX_RE = re.compile(r"^\s*\d{6,}\s*[-–]\s*")
#: The site repeats the serial in brackets at the end of a dropdown entry.
_SERIAL_SUFFIX_RE = re.compile(r"\s*\(\s*\d{6,}\s*\)\s*$")
_CODE_PREFIX_RE = re.compile(r"^\(\s*[A-Za-zČĆŠĐŽčćšđž0-9\- ]{1,12}\s*\)\s*")
_LEGACY_DIRECTION_RE = re.compile(r"\s*-\s*([IU])\s*$", re.IGNORECASE)

#: Dropdown entries that are chrome rather than devices.
_NAVIGATION_LABELS = frozenset(
    {"home", "device list", "devices", "user settings", "log out", "logout", "login", "menu", "settings"}
)


def normalize(text: str | None) -> str:
    """Fold accents, lowercase, and reduce punctuation to single spaces."""
    decomposed = unicodedata.normalize("NFKD", str(text or ""))
    stripped = "".join(ch for ch in decomposed if not unicodedata.combining(ch))
    collapsed = re.sub(r"[^a-zA-Z0-9]+", " ", stripped.lower())
    return collapsed.strip()


@dataclass(frozen=True)
class ExcelRow:
    """One writable row of the report template."""

    row: int
    lokacija: str
    vodomjer: str

    @property
    def key(self) -> tuple[str, str]:
        return (normalize(self.lokacija), normalize(self.vodomjer))

    @property
    def label(self) -> str:
        return f"{self.lokacija} / {self.vodomjer}" if self.vodomjer else self.lokacija


@dataclass(frozen=True)
class Station:
    """One scrape job: read ``uredjaj`` on the site, write it into the Excel row."""

    lokacija: str
    vodomjer: str
    uredjaj: str
    enabled: bool = True

    @property
    def key(self) -> tuple[str, str]:
        return (normalize(self.lokacija), normalize(self.vodomjer))

    @property
    def label(self) -> str:
        return f"{self.lokacija} / {self.vodomjer}" if self.vodomjer else self.lokacija

    def to_json(self) -> dict[str, Any]:
        payload: dict[str, Any] = {
            "lokacija": self.lokacija,
            "vodomjer": self.vodomjer,
            "uredjaj": self.uredjaj,
        }
        if not self.enabled:
            payload["enabled"] = False
        return payload

    @classmethod
    def from_json(cls, payload: dict[str, Any]) -> "Station":
        return cls(
            lokacija=str(payload.get("lokacija") or "").strip(),
            vodomjer=str(payload.get("vodomjer") or "").strip(),
            uredjaj=str(payload.get("uredjaj") or "").strip(),
            enabled=bool(payload.get("enabled", True)),
        )


@dataclass(frozen=True)
class Issue:
    """A problem found while checking stations against the template."""

    severity: str  # "error" | "warning"
    station: str
    message: str

    def to_json(self) -> dict[str, str]:
        return {"severity": self.severity, "station": self.station, "message": self.message}


# --------------------------------------------------------------------------- #
# Device labels
# --------------------------------------------------------------------------- #


def device_label(option_text: str, city_prefixes: Iterable[str] = DEFAULT_CITY_PREFIXES) -> str:
    """Reduce a raw site dropdown entry to the label stored in the registry.

    ``"358004092234384 - Herceg Novi - (R-BA) Bajer 1 - U (358004092234384)"``
    becomes ``"(R-BA) Bajer 1 - U"``.

    Applying this to a stored label too is what lets someone paste a dropdown
    entry straight off the site: the reduction is idempotent, so a label that
    was already written the short way comes back unchanged.
    """
    text = re.sub(r"\s+", " ", str(option_text or "")).strip()
    text = _SERIAL_PREFIX_RE.sub("", text)
    text = _SERIAL_SUFFIX_RE.sub("", text)
    for city in city_prefixes:
        # The city is the last prefix segment, so cut everything up to and
        # including it. That also drops placeholders the old file used in place
        # of a serial number.
        match = re.search(rf"(?:^|[-–]\s*){re.escape(city)}\s*[-–]\s*", text, flags=re.IGNORECASE)
        if match:
            text = text[match.end() :]
            break
    return text.strip()


def is_navigation_option(option_text: str) -> bool:
    return normalize(option_text) in _NAVIGATION_LABELS


def score_device_match(search: str, option_label_text: str) -> int:
    """Rank how well a site dropdown entry answers a stored device label.

    Higher is better, ``0`` means "not this one". The tiers are deliberately far
    apart so that a tie means genuine ambiguity rather than rounding.

    Both sides go through :func:`device_label` first, so a stored name pasted
    whole off the site -- serial, city and all -- is compared against the same
    reduction as the entry it came from instead of scoring zero.
    """
    wanted = normalize(device_label(search))
    found = normalize(device_label(option_label_text))
    if not wanted or not found:
        return 0
    if wanted == found:
        return 1000
    if found.endswith(f" {wanted}"):
        return 600
    if found.startswith(f"{wanted} "):
        return 400
    if f" {wanted} " in f" {found} ":
        return 300
    wanted_tokens = set(wanted.split())
    found_tokens = set(found.split())
    if wanted_tokens and wanted_tokens <= found_tokens:
        return 100 + len(wanted_tokens)
    return 0


def search_queries(uredjaj: str) -> list[str]:
    """Text to type into the site's search box, most specific first.

    The site filters on a literal substring of the full entry, so the stored
    label works as-is. The fallbacks cover sites that trip over punctuation.
    """
    queries: list[str] = []
    candidates = [
        uredjaj,
        _CODE_PREFIX_RE.sub("", uredjaj).strip(),
        re.sub(r"\s*[-–]\s*[IU]\s*$", "", _CODE_PREFIX_RE.sub("", uredjaj), flags=re.IGNORECASE).strip(),
    ]
    for candidate in candidates:
        candidate = re.sub(r"\s+", " ", candidate).strip()
        if candidate and candidate not in queries:
            queries.append(candidate)
    return queries


def shortest_unique_label(label: str, all_labels: Iterable[str]) -> str:
    """Drop the ``(R-BA)`` style prefix when doing so stays unambiguous."""
    short = _CODE_PREFIX_RE.sub("", label).strip()
    if not short or short == label:
        return label
    competitors = [other for other in all_labels if normalize(other) != normalize(label)]
    if any(score_device_match(short, other) >= 300 for other in competitors):
        return label
    return short


# --------------------------------------------------------------------------- #
# Excel template rows
# --------------------------------------------------------------------------- #


def load_excel_rows(template_path: Path) -> list[ExcelRow]:
    """Read every writable ``LOKACIJA`` / ``VODOMJER`` row from the template.

    ``LOKACIJA`` is only written on the first row of a group, so it is carried
    down until the next non-empty cell.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(template_path, data_only=True, read_only=True)
    try:
        for worksheet in workbook.worksheets:
            header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = {str(value).strip().upper(): index for index, value in enumerate(header) if value}
            if not {"LOKACIJA", "VODOMJER"} <= headers.keys():
                continue

            location_index = headers["LOKACIJA"]
            meter_index = headers["VODOMJER"]
            rows: list[ExcelRow] = []
            current_location = ""
            for offset, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                location = str(values[location_index]).strip() if _has(values, location_index) else ""
                meter = str(values[meter_index]).strip() if _has(values, meter_index) else ""
                if location:
                    current_location = location
                if not current_location or not meter:
                    continue
                rows.append(ExcelRow(row=offset, lokacija=current_location, vodomjer=meter))
            return rows
        raise RuntimeError(f"{template_path.name} has no sheet with LOKACIJA and VODOMJER headers.")
    finally:
        workbook.close()


def _has(values: tuple[Any, ...], index: int) -> bool:
    return index < len(values) and values[index] is not None


def index_excel_rows(rows: Iterable[ExcelRow]) -> dict[tuple[str, str], ExcelRow]:
    """Index rows by normalized ``(LOKACIJA, VODOMJER)``; first row wins a clash."""
    index: dict[tuple[str, str], ExcelRow] = {}
    for row in rows:
        index.setdefault(row.key, row)
    return index


# --------------------------------------------------------------------------- #
# Registry file
# --------------------------------------------------------------------------- #


def resolve_stations_path(explicit: str | Path | None, root: Path) -> Path:
    """Pick the registry file: an explicit path, then the new file, then legacy."""
    if explicit:
        return Path(explicit).expanduser()
    preferred = root / DEFAULT_STATIONS_FILE
    if preferred.exists():
        return preferred
    legacy = root / LEGACY_STATIONS_FILE
    if legacy.exists():
        return legacy
    return preferred


def load_stations(path: Path) -> list[Station]:
    """Read the registry, transparently upgrading the legacy serial-number file."""
    if not path.exists():
        return []
    raw = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(raw, dict) and "stations" in raw:
        return [Station.from_json(item) for item in raw.get("stations") or []]
    if isinstance(raw, dict):
        return migrate_legacy_mapping(raw)
    if isinstance(raw, list):
        return [Station.from_json(item) for item in raw]
    raise ValueError(f"{path.name} is not a station registry.")


def save_stations(path: Path, stations: Iterable[Station]) -> None:
    payload = {
        "version": SCHEMA_VERSION,
        "stations": [station.to_json() for station in stations],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def migrate_legacy_mapping(mapping: dict[str, Any]) -> list[Station]:
    """Convert ``{"REZERVOAR BAJER 1 - U": "3580... - Herceg Novi - ..."}``.

    The legacy key packed ``LOKACIJA`` and a ``- U`` / ``- I`` direction into one
    string; the direction becomes ``ULAZ`` / ``IZLAZ``. Keys that also carried a
    descriptive meter name cannot be split reliably, so they are left for
    :func:`reconcile_with_template` to resolve against the real template rows.
    """
    stations: list[Station] = []
    for key, value in mapping.items():
        name = str(key).strip()
        if not name:
            continue
        direction = _LEGACY_DIRECTION_RE.search(name)
        lokacija = _LEGACY_DIRECTION_RE.sub("", name).strip()
        vodomjer = ""
        if direction:
            vodomjer = "ULAZ" if direction.group(1).upper() == "U" else "IZLAZ"
        stations.append(
            Station(lokacija=lokacija, vodomjer=vodomjer, uredjaj=device_label(str(value)))
        )
    return stations


def reconcile_with_template(stations: Iterable[Station], rows: Iterable[ExcelRow]) -> list[Station]:
    """Snap each station onto a real template row.

    Exact ``(LOKACIJA, VODOMJER)`` hits are kept as-is. Anything else is matched
    against the unclaimed rows by comparing the station's full text with the
    row's ``LOKACIJA VODOMJER`` text, which is what resolves legacy keys such as
    ``REZERVOAR BAJER 2 - IZLAZ ZA ČELA``.
    """
    row_list = list(rows)
    index = index_excel_rows(row_list)
    resolved: list[Station] = []
    claimed: set[int] = set()
    pending: list[Station] = []

    for station in stations:
        row = index.get(station.key)
        if row and row.row not in claimed:
            claimed.add(row.row)
            resolved.append(replace(station, lokacija=row.lokacija, vodomjer=row.vodomjer))
        else:
            pending.append(station)
            resolved.append(station)

    for station in pending:
        wanted = normalize(f"{station.lokacija} {station.vodomjer}")
        wanted_tokens = set(wanted.split())
        best: tuple[int, ExcelRow] | None = None
        for row in row_list:
            if row.row in claimed:
                continue
            candidate = normalize(f"{row.lokacija} {row.vodomjer}")
            score = max(score_device_match(candidate, wanted), score_device_match(wanted, candidate))
            if not score:
                # Neither text contains the other, which is normal when the two
                # spellings differ mid-string ("PUNJENJE REZERVOARA PODI" vs
                # "ULAZ REZERVOARA PODI"). Fall back to shared-word overlap.
                candidate_tokens = set(candidate.split())
                shared = wanted_tokens & candidate_tokens
                union = wanted_tokens | candidate_tokens
                if union and len(shared) / len(union) >= 0.5:
                    score = 10 + len(shared)
            if score and (best is None or score > best[0]):
                best = (score, row)
        if best:
            claimed.add(best[1].row)
            index_in_result = resolved.index(station)
            resolved[index_in_result] = replace(
                station, lokacija=best[1].lokacija, vodomjer=best[1].vodomjer
            )
    return resolved


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #


def validate(stations: Iterable[Station], rows: Iterable[ExcelRow]) -> list[Issue]:
    """Report everything that would make a run write to the wrong place.

    Errors block a station from running; warnings are informational.
    """
    station_list = list(stations)
    row_list = list(rows)
    index = index_excel_rows(row_list)
    issues: list[Issue] = []

    seen_rows: dict[tuple[str, str], Station] = {}
    seen_devices: dict[str, Station] = {}

    for station in station_list:
        label = station.label or "(prazan unos)"
        if not station.lokacija:
            issues.append(Issue("error", label, "LOKACIJA je obavezna."))
        if not station.vodomjer:
            issues.append(Issue("error", label, "VODOMJER je obavezan."))
        if not station.uredjaj:
            issues.append(Issue("error", label, "Naziv uređaja na sajtu je obavezan."))

        if station.lokacija and station.vodomjer and station.key not in index:
            issues.append(
                Issue(
                    "error",
                    label,
                    "Nema reda u template-u sa ovom kombinacijom LOKACIJA + VODOMJER.",
                )
            )

        previous = seen_rows.get(station.key)
        if previous is not None:
            issues.append(Issue("error", label, "Dva unosa pišu u isti red template-a."))
        else:
            seen_rows[station.key] = station

        device_key = normalize(station.uredjaj)
        if device_key:
            duplicate = seen_devices.get(device_key)
            if duplicate is not None:
                issues.append(
                    Issue(
                        "warning",
                        label,
                        f"Isti uređaj se čita i za '{duplicate.label}'. To je u redu samo ako je namjerno.",
                    )
                )
            else:
                seen_devices[device_key] = station

    covered = {station.key for station in station_list}
    for row in row_list:
        if row.key not in covered:
            issues.append(Issue("warning", row.label, f"Red {row.row} u template-u nema unos i ostaće prazan."))

    return issues


def ambiguous_device_labels(stations: Iterable[Station]) -> list[Issue]:
    """Flag stored labels that would match more than one another stored label."""
    station_list = [station for station in stations if station.uredjaj]
    issues: list[Issue] = []
    for station in station_list:
        rivals = [
            other.uredjaj
            for other in station_list
            if other is not station
            and normalize(other.uredjaj) != normalize(station.uredjaj)
            and score_device_match(station.uredjaj, other.uredjaj) >= 300
        ]
        if rivals:
            issues.append(
                Issue(
                    "warning",
                    station.label,
                    f"Naziv '{station.uredjaj}' odgovara i uređaju '{rivals[0]}'. Dodaj prefiks da bude jedinstven.",
                )
            )
    return issues
